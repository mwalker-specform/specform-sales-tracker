#!/usr/bin/env python3
"""
RMAX Quote Tracker — FastAPI backend
Run: uvicorn backend:app --reload --port 8000
"""
import os, sqlite3, json, glob as _glob, shutil, tempfile, io, time, csv, zipfile, asyncio
from collections import defaultdict
from datetime import datetime, timedelta, date
from contextlib import contextmanager
from typing import Optional

# ── Date parsing helper ───────────────────────────────────────────────────────
_DATE_FMTS = [
    '%Y-%m-%d',   # 2026-07-09
    '%m/%d/%Y',   # 7/9/2026
    '%m/%d/%y',   # 7/9/26
    '%b %d, %Y',  # Jul 13, 2026
    '%B %d, %Y',  # July 13, 2026
    '%b. %d, %Y', # Jul. 13, 2026
]

def parse_date(s: str):
    """Return datetime or None for any of our known date string formats."""
    if not s:
        return None
    s = s.strip()
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def normalize_date(s: str) -> str:
    """Convert any supported date string to YYYY-MM-DD; return original if unparseable."""
    if not s:
        return s
    dt = parse_date(s)
    return dt.strftime('%Y-%m-%d') if dt else s

from fastapi import FastAPI, HTTPException, Query, UploadFile, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import bcrypt as _bcrypt
from jose import JWTError, jwt

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# DATA_DIR can be overridden via env var so the DB lives on a persistent Railway volume.
# In Railway: set DATA_DIR=/data and mount a volume at /data.
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)
BACKUP_DIR = os.path.join(DATA_DIR, 'backups')
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH  = os.path.join(DATA_DIR, "quotes.db")
STATIC   = os.path.join(BASE_DIR, "static")

# ── Auth config ───────────────────────────────────────────────────────────────
JWT_SECRET      = os.environ.get('JWT_SECRET', 'specform-salespartner-change-in-production')
JWT_ALGORITHM   = 'HS256'
JWT_EXPIRE_DAYS = 60

def _hash_pw(password: str) -> str:
    return _bcrypt.hashpw(password.encode('utf-8'), _bcrypt.gensalt()).decode('utf-8')

def _verify_pw(password: str, hashed: str) -> bool:
    try:
        return _bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        return False

# ── DB helpers ────────────────────────────────────────────────────────────────
@contextmanager
def get_db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        yield con
        con.commit()
    finally:
        con.close()

def init_db():
    with get_db() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS quotes (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            status             TEXT,
            date_received      TEXT,
            date_quoted        TEXT,
            sent_to            TEXT,
            subject            TEXT,
            job_name           TEXT,
            customer           TEXT,
            location           TEXT,
            product            TEXT,
            price              TEXT,
            quantities         TEXT,
            amount             REAL,
            close_date         TEXT,
            est_freight        TEXT,
            lead_time          TEXT,
            notes              TEXT,
            add_to_salesforce  INTEGER DEFAULT 0,
            completed          INTEGER DEFAULT 0,
            created_at         TEXT DEFAULT (datetime('now')),
            updated_at         TEXT DEFAULT (datetime('now'))
        )""")
        # Add columns to existing databases that predate this schema
        for col, definition in [('add_to_salesforce', 'INTEGER DEFAULT 0'),
                                 ('completed',         'INTEGER DEFAULT 0'),
                                 ('completed_date',    'TEXT'),
                                 ('region',            'TEXT'),
                                 ('deleted',           'INTEGER DEFAULT 0'),
                                 ('company_id',        'INTEGER'),
                                 ('architect_id',      'INTEGER')]:
            try:
                con.execute(f"ALTER TABLE quotes ADD COLUMN {col} {definition}")
            except Exception:
                pass  # Column already exists

# ── Models ────────────────────────────────────────────────────────────────────
STATUS_OPTIONS = ["Won", "Lost", "Not Awarded", "Duplicate", "Verbal", "Unlikely"]

class QuoteIn(BaseModel):
    status:             Optional[str] = None
    date_received:      Optional[str] = None
    date_quoted:        Optional[str] = None
    sent_to:            Optional[str] = None
    subject:            Optional[str] = None
    job_name:           Optional[str] = None
    customer:           Optional[str] = None
    location:           Optional[str] = None
    product:            Optional[str] = None
    price:              Optional[str] = None
    quantities:         Optional[str] = None
    amount:             Optional[float] = None
    close_date:         Optional[str] = None
    est_freight:        Optional[str] = None
    lead_time:          Optional[str] = None
    notes:              Optional[str] = None
    region:             Optional[str] = None
    add_to_salesforce:  Optional[int] = 0
    completed:          Optional[int] = 0
    company_id:         Optional[int] = None
    architect_id:       Optional[int] = None

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="RMAX Quote Tracker")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Auth middleware ────────────────────────────────────────────────────────────
_UNPROTECTED = {'/api/auth/login'}

@app.middleware('http')
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    # Pass through: static files, SPA root, OPTIONS preflight, login endpoint
    if (not path.startswith('/api/')
            or path in _UNPROTECTED
            or request.method == 'OPTIONS'):
        return await call_next(request)
    auth_header = request.headers.get('Authorization', '')
    token = auth_header[7:] if auth_header.startswith('Bearer ') else ''
    if not token:
        return JSONResponse({'detail': 'Not authenticated'}, status_code=401)
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        # Confirm the account is still active in the database
        with get_db() as con:
            row = con.execute(
                "SELECT is_active FROM users WHERE LOWER(email)=?",
                (payload.get('sub', '').lower(),)
            ).fetchone()
        if not row or not row['is_active']:
            return JSONResponse({'detail': 'Account deactivated'}, status_code=401)
        request.state.user = payload
    except JWTError:
        return JSONResponse({'detail': 'Invalid or expired token'}, status_code=401)
    return await call_next(request)

# ── Customer Prospects ────────────────────────────────────────────────────────────────────────────────
def init_prospects_db():
    with get_db() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS prospects (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name TEXT,
                contact_name TEXT,
                title        TEXT,
                phone        TEXT,
                email        TEXT,
                region       TEXT,
                status       TEXT DEFAULT 'Cold',
                notes        TEXT,
                created_at   TEXT DEFAULT (datetime('now'))
            )
        """)

class ProspectIn(BaseModel):
    company_name: str = ""
    contact_name: str = ""
    title:        str = ""
    phone:        str = ""
    email:        str = ""
    region:       str = ""
    status:       str = "Cold"
    notes:        str = ""

@app.get('/api/prospects')
def list_prospects(request: Request):
    _require_auth(request)
    with get_db() as con:
        rows = con.execute("SELECT * FROM prospects ORDER BY company_name").fetchall()
    return [dict(r) for r in rows]

@app.post('/api/prospects')
def create_prospect(p: ProspectIn, request: Request):
    _require_auth(request)
    with get_db() as con:
        cur = con.execute(
            "INSERT INTO prospects (company_name,contact_name,title,phone,email,region,status,notes) VALUES (?,?,?,?,?,?,?,?)",
            (p.company_name,p.contact_name,p.title,p.phone,p.email,p.region,p.status,p.notes)
        )
        return {"id": cur.lastrowid}

@app.put('/api/prospects/{pid}')
def update_prospect(pid: int, p: ProspectIn, request: Request):
    _require_auth(request)
    with get_db() as con:
        con.execute(
            "UPDATE prospects SET company_name=?,contact_name=?,title=?,phone=?,email=?,region=?,status=?,notes=? WHERE id=?",
            (p.company_name,p.contact_name,p.title,p.phone,p.email,p.region,p.status,p.notes,pid)
        )
    return {"ok": True}

@app.delete('/api/prospects/{pid}')
def delete_prospect(pid: int, request: Request):
    _require_admin(request)
    with get_db() as con:
        con.execute("DELETE FROM prospects WHERE id=?", (pid,))
    return {"ok": True}

# ── Backup ───────────────────────────────────────────────────────────────────
def _make_backup_zip() -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        with get_db() as con:
            tables = con.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()
            for row in tables:
                table = row['name']
                try:
                    rows = con.execute(f"SELECT * FROM [{table}]").fetchall()
                    if not rows: continue
                    csv_buf = io.StringIO()
                    writer = csv.DictWriter(csv_buf, fieldnames=rows[0].keys())
                    writer.writeheader()
                    writer.writerows([dict(r) for r in rows])
                    zf.writestr(f"{table}.csv", csv_buf.getvalue())
                except Exception: pass
    buf.seek(0)
    return buf.read()

def _send_backup_alert(error_msg: str):
    """Send an email alert when a backup fails."""
    try:
        import msal
        authority = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}"
        app_client = msal.ConfidentialClientApplication(
            GRAPH_CLIENT_ID, authority=authority, client_credential=GRAPH_CLIENT_SECRET
        )
        result = app_client.acquire_token_for_client(scopes=['https://graph.microsoft.com/.default'])
        if 'access_token' not in result:
            return
        admin_email = os.environ.get('BACKUP_ALERT_EMAIL', GRAPH_USER)
        requests.post(
            f'https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/sendMail',
            headers={'Authorization': f'Bearer {result["access_token"]}', 'Content-Type': 'application/json'},
            json={'message': {'subject': '[SPECFORM] Backup Failed', 'body': {'contentType': 'Text', 'content': f'Automated backup failed on {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}:\n\n{error_msg}'}, 'toRecipients': [{'emailAddress': {'address': admin_email}}]}},
            timeout=30
        )
    except Exception as alert_err:
        print(f'[backup] alert email failed: {alert_err}')

def _upload_to_onedrive(zip_bytes: bytes, filename: str):
    """Upload a backup zip to OneDrive via Microsoft Graph API."""
    try:
        import msal
    except ImportError:
        raise Exception("msal not installed")
    authority = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}"
    app_client = msal.ConfidentialClientApplication(
        GRAPH_CLIENT_ID, authority=authority, client_credential=GRAPH_CLIENT_SECRET
    )
    result = app_client.acquire_token_for_client(scopes=['https://graph.microsoft.com/.default'])
    if 'access_token' not in result:
        raise Exception(f"MSAL token error: {result.get('error_description','unknown')}")
    token = result['access_token']
    folder = 'specform-backups'
    url = f'https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/drive/root:/{folder}/{filename}:/content'
    resp = requests.put(url, headers={
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/octet-stream'
    }, data=zip_bytes, timeout=60)
    resp.raise_for_status()

async def _daily_backup_task():
    while True:
        now = datetime.utcnow()
        tomorrow = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        await asyncio.sleep((tomorrow - now).total_seconds())
        try:
            data = _make_backup_zip()
            date_str = datetime.utcnow().strftime('%Y-%m-%d')
            with open(os.path.join(BACKUP_DIR, f'backup-{date_str}.zip'), 'wb') as fh:
                fh.write(data)
            try:
                _upload_to_onedrive(data, f'backup-{date_str}.zip')
                print(f'[backup] uploaded to OneDrive: backup-{date_str}.zip')
            except Exception as oe:
                print(f'[backup] OneDrive upload failed: {oe}')
            old = sorted(_glob.glob(os.path.join(BACKUP_DIR, 'backup-*.zip')))[:-7]
            for p in old: os.remove(p)
        except Exception as e:
            print(f'[backup] daily backup failed: {e}')
            _send_backup_alert(str(e))

@app.get('/api/admin/backup')
def admin_download_backup(request: Request):
    _require_admin(request)
    date_str = datetime.utcnow().strftime('%Y-%m-%d')
    zip_bytes = _make_backup_zip()
    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type='application/zip',
        headers={'Content-Disposition': f'attachment; filename="specform-backup-{date_str}.zip"'}
    )

@app.on_event("startup")
async def startup():
    init_db()
    init_hydrotech_db()
    init_glassworks_db()
    init_lam_db()
    init_contacts_db()
    init_prospects_db()
    init_users_db()
    migrate_dates()
    os.makedirs(BACKUP_DIR, exist_ok=True)
    asyncio.create_task(_daily_backup_task())

def migrate_dates():
    """One-time migration: normalize any non-ISO date strings in existing rows."""
    date_cols = ['date_received', 'date_quoted', 'close_date']
    for table in ('quotes', 'hydrotech_quotes'):
        try:
            with get_db() as con:
                rows = con.execute(f"SELECT id, {', '.join(date_cols)} FROM {table}").fetchall()
                for row in rows:
                    updates = {}
                    for col in date_cols:
                        val = row[col]
                        if val and not val[:4].isdigit():  # already ISO if starts with 4 digits
                            normalized = normalize_date(val)
                            if normalized != val:
                                updates[col] = normalized
                    if updates:
                        sets = ', '.join(f"{c}=?" for c in updates)
                        con.execute(f"UPDATE {table} SET {sets} WHERE id=?",
                                    list(updates.values()) + [row['id']])
        except Exception:
            pass

# ── Quote endpoints ───────────────────────────────────────────────────────────
@app.get("/api/quotes")
def list_quotes(
    status:   Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    search:   Optional[str] = Query(None),
):
    with get_db() as con:
        sql = "SELECT * FROM quotes WHERE (deleted IS NULL OR deleted=0)"
        params = []
        if status and status != "All":
            sql += " AND status = ?"
            params.append(status)
        if location and location != "All":
            sql += " AND location = ?"
            params.append(location)
        if search:
            sql += " AND (subject LIKE ? OR sent_to LIKE ? OR job_name LIKE ? OR customer LIKE ? OR location LIKE ?)"
            s = f"%{search}%"
            params += [s, s, s, s, s]
        sql += " ORDER BY date_received DESC"
        rows = con.execute(sql, params).fetchall()
        quotes = [dict(r) for r in rows]
        if quotes:
            ids = [q['id'] for q in quotes]
            file_rows = con.execute(
                f"SELECT id, quote_id, file_filename FROM glassworks_quote_files WHERE quote_id IN ({','.join('?'*len(ids))})",
                ids
            ).fetchall()
            file_map = {}
            for fr in file_rows:
                file_map.setdefault(fr['quote_id'], []).append({'id': fr['id'], 'filename': fr['file_filename']})
            for q in quotes:
                q['files'] = file_map.get(q['id'], [])
        return quotes

@app.get("/api/quotes/{quote_id}")
def get_quote(quote_id: int):
    with get_db() as con:
        row = con.execute("SELECT * FROM quotes WHERE id=? AND (deleted IS NULL OR deleted=0)", (quote_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Quote not found")
        return dict(row)

@app.post("/api/quotes", status_code=201)
def create_quote(q: QuoteIn):
    with get_db() as con:
        cur = con.execute("""
        INSERT INTO quotes (status,date_received,date_quoted,sent_to,subject,job_name,
            customer,location,product,price,quantities,amount,close_date,est_freight,lead_time,notes,
            region,add_to_salesforce,completed,company_id,architect_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (q.status,normalize_date(q.date_received),normalize_date(q.date_quoted),
              q.sent_to,q.subject,q.job_name,
              q.customer,q.location,q.product,q.price,q.quantities,q.amount,
              normalize_date(q.close_date),q.est_freight,q.lead_time,q.notes,
              q.region,q.add_to_salesforce,q.completed,q.company_id,q.architect_id))
        return {"id": cur.lastrowid}

@app.put("/api/quotes/{quote_id}")
def update_quote(quote_id: int, q: QuoteIn):
    with get_db() as con:
        existing = con.execute("SELECT completed, completed_date FROM quotes WHERE id=?", (quote_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Quote not found")
        if q.completed and not existing['completed_date']:
            completed_date = date.today().strftime('%Y-%m-%d')
        elif not q.completed:
            completed_date = None
        else:
            completed_date = existing['completed_date']
        con.execute("""
        UPDATE quotes SET status=?,date_received=?,date_quoted=?,sent_to=?,subject=?,
            job_name=?,customer=?,location=?,product=?,price=?,quantities=?,amount=?,
            close_date=?,est_freight=?,lead_time=?,notes=?,
            region=?,add_to_salesforce=?,completed=?,completed_date=?,company_id=?,architect_id=?,updated_at=datetime('now')
        WHERE id=?
        """, (q.status,normalize_date(q.date_received),normalize_date(q.date_quoted),
              q.sent_to,q.subject,q.job_name,
              q.customer,q.location,q.product,q.price,q.quantities,q.amount,
              normalize_date(q.close_date),q.est_freight,q.lead_time,q.notes,
              q.region,q.add_to_salesforce,q.completed,completed_date,q.company_id,q.architect_id,quote_id))
        return {"ok": True}

@app.delete("/api/quotes/{quote_id}")
def delete_quote(quote_id: int):
    with get_db() as con:
        # Soft delete — keeps row so sync won't re-import the same email
        con.execute("UPDATE quotes SET deleted=1 WHERE id=?", (quote_id,))
        return {"ok": True}

@app.get("/api/quotes-export")
def export_quotes():
    """Export RMAX quotes grouped by month → region, styled like Company Accounts."""
    import io
    import openpyxl
    from collections import defaultdict
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    thin     = Side(style="thin", color="D1D5DB")
    BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT     = Alignment(horizontal="left", vertical="center")
    INDENTED = Alignment(horizontal="left", vertical="center", indent=1)

    HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT    = Font(bold=True, color="FFFFFF", size=11)
    MONTH_FILL  = PatternFill("solid", fgColor="1F4E79")
    MONTH_FONT  = Font(bold=True, color="FFFFFF", size=12)
    REGION_FILL = PatternFill("solid", fgColor="DBEAFE")
    REGION_FONT = Font(bold=True, color="1E3A5F", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="F4F6F8")
    STATUS_FILLS = {
        "Won":         PatternFill("solid", fgColor="D9EAD3"),
        "Lost":        PatternFill("solid", fgColor="F4CCCC"),
        "Verbal":      PatternFill("solid", fgColor="FFF2CC"),
        "Not Awarded": PatternFill("solid", fgColor="FFE4CC"),
        "Unlikely":    PatternFill("solid", fgColor="EDE9FE"),
        "Duplicate":   PatternFill("solid", fgColor="F3F4F6"),
    }
    MONEY_FMT    = "$#,##0.00"
    REGION_ORDER = ["Central Texas", "Southeast Texas", "The Valley"]
    NCOLS        = 13

    HEADERS = [
        ("Date Received", 14), ("Status", 12), ("Job Name", 28), ("Customer", 22),
        ("Location", 18), ("Sent To", 20), ("Product", 16), ("Price", 12),
        ("Quantities", 14), ("Amount ($)", 14), ("Close Date", 14),
        ("Est. Freight", 14), ("Notes", 35),
    ]

    with get_db() as con:
        rows = con.execute("""
            SELECT date_received, status, job_name, subject, customer, location,
                   region, sent_to, product, price, quantities, amount,
                   close_date, est_freight, lead_time, notes
            FROM quotes WHERE (deleted IS NULL OR deleted=0)
            ORDER BY date_received DESC
        """).fetchall()

    # Group: month (YYYY-MM) → region → [quotes]
    month_region = defaultdict(lambda: defaultdict(list))
    no_date_region = defaultdict(list)

    for q in rows:
        dr = q["date_received"] or ""
        dt = parse_date(dr)
        ym = dt.strftime("%Y-%m") if dt else None
        region = (q["region"] or "").strip() or "(No Region)"
        if ym:
            month_region[ym][region].append(q)
        else:
            no_date_region[region].append(q)

    sorted_months = sorted(month_region.keys(), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RMAX Quotes"

    # ── Column header row ──────────────────────────────────────────────────────
    for col, (label, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CTR, BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22

    cur_row = 2

    def region_sort_key(r):
        try:
            return (REGION_ORDER.index(r), r)
        except ValueError:
            return (len(REGION_ORDER), r)

    def write_group(ym, region_dict):
        nonlocal cur_row

        # Month label
        if ym and ym != "(No Date)":
            try:
                month_label = datetime.strptime(ym + "-01", "%Y-%m-%d").strftime("%B %Y")
            except Exception:
                month_label = ym
        else:
            month_label = "(No Date)"

        all_qs     = [q for qs in region_dict.values() for q in qs]
        month_tot  = sum(q["amount"] or 0 for q in all_qs)

        # Month header row
        for col in range(1, NCOLS + 1):
            cell = ws.cell(row=cur_row, column=col)
            if col == 1:
                cell.value, cell.font, cell.alignment = month_label, MONTH_FONT, LEFT
            elif col == NCOLS:
                cell.value = f"{len(all_qs)} quote{'s' if len(all_qs) != 1 else ''} · ${month_tot:,.0f}"
                cell.font      = Font(bold=True, color="BDD7EE", size=10)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill, cell.border = MONTH_FILL, BORDER
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        for region in sorted(region_dict.keys(), key=region_sort_key):
            q_list     = region_dict[region]
            region_tot = sum(q["amount"] or 0 for q in q_list)

            # Region sub-header row
            for col in range(1, NCOLS + 1):
                cell = ws.cell(row=cur_row, column=col)
                if col == 1:
                    cell.value, cell.font, cell.alignment = region, REGION_FONT, INDENTED
                elif col == NCOLS:
                    cell.value = f"{len(q_list)} quote{'s' if len(q_list) != 1 else ''} · ${region_tot:,.0f}"
                    cell.font      = Font(bold=True, color="1E3A5F", size=9)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.fill, cell.border = REGION_FILL, BORDER
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

            # Quote data rows
            for idx, q in enumerate(q_list):
                status   = q["status"] or ""
                row_fill = STATUS_FILLS.get(status, (ALT_FILL if idx % 2 == 1 else PatternFill()))
                row_font = Font(size=9, bold=(status == "Won"))

                vals = [
                    q["date_received"],
                    q["status"],
                    q["job_name"] or q["subject"] or "",
                    q["customer"],
                    q["location"],
                    q["sent_to"],
                    q["product"],
                    q["price"],
                    q["quantities"],
                    q["amount"],
                    q["close_date"],
                    q["est_freight"],
                    q["notes"],
                ]

                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=cur_row, column=col, value=v)
                    cell.fill, cell.font, cell.border = row_fill, row_font, BORDER
                    cell.alignment = Alignment(wrap_text=(col == NCOLS), vertical="top")
                    if col == 10 and v is not None:
                        cell.number_format = MONEY_FMT
                    if col == 2:
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                ws.row_dimensions[cur_row].height = 14
                cur_row += 1

    for ym in sorted_months:
        write_group(ym, month_region[ym])

    if no_date_region:
        write_group("(No Date)", no_date_region)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"RMAX_Quotes_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ── Sync endpoint — pulls directly from Outlook "RMAX Quotes" folder ──────────
import re as _sync_re

def _strip_tags(html: str) -> str:
    """Minimal HTML→text for use before _strip_html is defined."""
    html = _sync_re.sub(r'<style[^>]*>.*?</style>', '', html, flags=_sync_re.DOTALL|_sync_re.IGNORECASE)
    html = _sync_re.sub(r'<script[^>]*>.*?</script>', '', html, flags=_sync_re.DOTALL|_sync_re.IGNORECASE)
    html = _sync_re.sub(r'<br\s*/?>', '\n', html, flags=_sync_re.IGNORECASE)
    html = _sync_re.sub(r'</?(?:p|div|tr|td|li|h\d)[^>]*>', '\n', html, flags=_sync_re.IGNORECASE)
    html = _sync_re.sub(r'<[^>]+>', '', html)
    return (html.replace('&nbsp;', ' ').replace('&amp;', '&')
                .replace('&lt;', '<').replace('&gt;', '>').replace('\r', ''))

def _parse_quote_email(subject: str, body_html: str, content_type: str) -> dict:
    """
    Parse a quote-request email whose body follows the standard template:
        Job Name: XYZ
        Customer: XYZ
        Location: XYZ
        Product: XYZ
        Price: XYZ
        Quantities: XYZ
        Estimated Freight: XYZ
        Lead time: XYZ
    """
    body = _strip_tags(body_html) if content_type == "html" else body_html
    body = _sync_re.sub(r'\n{3,}', '\n\n', body).strip()

    def _field(label: str) -> str:
        """Extract the value after 'Label:' on the same line."""
        m = _sync_re.search(
            rf'^\s*{_sync_re.escape(label)}\s*:\s*(.+)',
            body, _sync_re.IGNORECASE | _sync_re.MULTILINE
        )
        return m.group(1).strip() if m else ''

    fields = {}

    job      = _field('Job Name')
    customer = _field('Customer')
    location = _field('Location')
    product  = _field('Product')
    price    = _field('Price')
    qty      = _field('Quantities')
    freight  = _field('Estimated Freight')
    lead     = _field('Lead time') or _field('Lead Time')

    # Fallback: derive job name from subject if body label missing
    if not job and subject:
        job = _sync_re.sub(
            r'^(re:|fw:|fwd:|(?:rmax|hydrotech|glassworks)\s+quotes?\s*[-–:]\s*)+',
            '', subject, flags=_sync_re.IGNORECASE
        ).strip()

    if job:      fields['job_name']    = job
    if customer: fields['customer']    = customer
    if location: fields['location']    = location
    if product:  fields['product']     = product
    if price:    fields['price']       = price
    if qty:      fields['quantities']  = qty
    if freight:  fields['est_freight'] = freight
    if lead:     fields['lead_time']   = lead

    # Amount: price_per_sf × sf_per_bundle × bundles
    # Price field example:  "$1.100/sf through Q3, 2026"
    # Qty field example:    "48 pcs/bundle (1536 sf/bundle) – 52 bundles"
    try:
        price_sf_m   = _sync_re.search(r'\$\s*([\d,]+(?:\.\d+)?)\s*/\s*sf', price or '', _sync_re.IGNORECASE)
        sf_bundle_m  = _sync_re.search(r'([\d,]+)\s*sf\s*/\s*bundle', qty or '', _sync_re.IGNORECASE)
        bundles_m    = _sync_re.search(r'(\d[\d,]*)\s*bundles?\b', qty or '', _sync_re.IGNORECASE)
        if price_sf_m and sf_bundle_m and bundles_m:
            price_per_sf  = float(price_sf_m.group(1).replace(',', ''))
            sf_per_bundle = float(sf_bundle_m.group(1).replace(',', ''))
            num_bundles   = float(bundles_m.group(1).replace(',', ''))
            fields['amount'] = round(price_per_sf * sf_per_bundle * num_bundles, 2)
    except (AttributeError, ValueError):
        pass

    return fields


@app.post("/api/sync")
def sync_from_outlook():
    """
    Pull new emails from the 'RMAX Quotes' Outlook folder via Graph API,
    parse each email body to pre-populate quote fields, and insert any
    that are not already in the database.
    Falls back to sync_pending.json if Graph is not configured.
    """
    # ── Primary path: Graph API ───────────────────────────────────────────────
    if GRAPH_CLIENT_SECRET:
        try:
            import requests as _req
        except ImportError:
            return {"inserted": 0, "skipped": 0, "message": "requests package not installed"}
        try:
            token = _get_graph_token()
        except Exception as e:
            return {"inserted": 0, "skipped": 0, "message": f"Auth failed: {e}"}

        hdrs = {"Authorization": f"Bearer {token}"}

        # ── Find "RMAX Quotes" folder ─────────────────────────────────────────
        # List all top-level and Inbox child folders, match by name
        folder_id = None
        for list_url in [
            f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders?$select=id,displayName&$top=100",
            f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders/Inbox/childFolders?$select=id,displayName&$top=100",
        ]:
            try:
                r = _req.get(list_url, headers=hdrs, timeout=20)
                if r.status_code == 200:
                    for f in r.json().get("value", []):
                        if (f.get("displayName") or "").strip().lower() == "rmax quotes":
                            folder_id = f["id"]
                            break
                if folder_id:
                    break
            except Exception:
                pass

        if not folder_id:
            return {"inserted": 0, "skipped": 0,
                    "message": "Could not find 'RMAX Quotes' folder in Outlook — check folder name"}

        # ── Fetch messages with body ──────────────────────────────────────────
        msgs = []
        next_url = (f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}"
                    f"/mailFolders/{folder_id}/messages"
                    f"?$select=from,toRecipients,subject,receivedDateTime,body&$top=50")
        while next_url and len(msgs) < 200:
            try:
                r = _req.get(next_url, headers=hdrs, timeout=30)
                if r.status_code != 200:
                    return {"inserted": 0, "skipped": 0,
                            "message": f"Graph API error {r.status_code}: {r.text[:200]}"}
                data = r.json()
                msgs.extend(data.get("value", []))
                next_url = data.get("@odata.nextLink")
            except Exception as e:
                return {"inserted": 0, "skipped": 0, "message": f"Fetch error: {e}"}

        # ── Insert / update quotes ────────────────────────────────────────────
        inserted = updated = skipped = 0
        with get_db() as con:
            for msg in msgs:
                # "Sent To" = first recipient (who the quote was sent to)
                to_recipients = msg.get("toRecipients", [])
                if to_recipients:
                    first_to = to_recipients[0].get("emailAddress", {})
                    sent_to  = (first_to.get("name") or first_to.get("address") or "").strip()
                else:
                    sent_to  = ""

                subject       = (msg.get("subject") or "").strip()
                received_raw  = msg.get("receivedDateTime", "")
                date_received = received_raw[:10] if received_raw else ""

                # Dedup check — if already exists, fix sent_to if it was wrong
                exists = con.execute(
                    "SELECT id, sent_to FROM quotes WHERE subject=? AND date_received=?",
                    (subject, date_received)
                ).fetchone()
                if exists:
                    # Only fill in sent_to if blank — never overwrite a user-edited value
                    if sent_to and not exists["sent_to"]:
                        con.execute("UPDATE quotes SET sent_to=? WHERE id=?",
                                    (sent_to, exists["id"]))
                        updated += 1
                    else:
                        skipped += 1
                    continue

                # Parse body for pre-populated fields
                body_content = msg.get("body", {}).get("content", "")
                content_type = msg.get("body", {}).get("contentType", "text")
                parsed = _parse_quote_email(subject, body_content, content_type)

                con.execute("""
                INSERT INTO quotes
                    (date_received, sent_to, subject, job_name, customer, location,
                     product, price, quantities, amount, est_freight, lead_time)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    date_received, sent_to, subject,
                    parsed.get('job_name'), parsed.get('customer'), parsed.get('location'),
                    parsed.get('product'), parsed.get('price'), parsed.get('quantities'),
                    parsed.get('amount'), parsed.get('est_freight'), parsed.get('lead_time'),
                ))
                inserted += 1

        parts = []
        if inserted: parts.append(f"{inserted} new quote{'' if inserted==1 else 's'} imported")
        if updated:  parts.append(f"{updated} updated")
        msg_text = "✓ " + ", ".join(parts) if parts else "✓ Already up to date"
        return {"inserted": inserted, "skipped": skipped, "message": msg_text}

    # ── Fallback: sync_pending.json (legacy path) ─────────────────────────────
    import json as _json
    pending_path = os.path.join(BASE_DIR, "sync_pending.json")
    if not os.path.exists(pending_path):
        return {"inserted": 0, "skipped": 0, "message": "Graph API not configured and no pending quotes file"}

    with open(pending_path, "r", encoding="utf-8") as f:
        records = _json.load(f)

    inserted = skipped = 0
    with get_db() as con:
        for r in records:
            exists = con.execute(
                "SELECT id FROM quotes WHERE sent_to=? AND subject=? AND date_received=?",
                (r.get("sent_to"), r.get("subject"), r.get("date_received"))
            ).fetchone()
            if exists:
                skipped += 1
                continue
            con.execute("""
            INSERT INTO quotes
                (status, date_received, date_quoted, sent_to, subject, job_name,
                 customer, location, product, price, quantities, amount,
                 close_date, est_freight, lead_time, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                r.get("status"), r.get("date_received"), r.get("date_quoted"),
                r.get("sent_to"), r.get("subject"), r.get("job_name"),
                r.get("customer"), r.get("location"), r.get("product"),
                r.get("price"), r.get("quantities"), r.get("amount"),
                r.get("close_date"), r.get("est_freight"), r.get("lead_time"),
                r.get("notes"),
            ))
            inserted += 1

    with open(pending_path, "w", encoding="utf-8") as f:
        _json.dump([], f)

    return {"inserted": inserted, "skipped": skipped,
            "message": f"Imported {inserted} new quote(s), {skipped} duplicate(s) skipped"}

# ── Sync debug endpoint ───────────────────────────────────────────────────────
@app.get("/api/sync-debug")
def sync_debug():
    """Diagnose Graph API connectivity and RMAX Quotes folder lookup."""
    if not GRAPH_CLIENT_SECRET:
        return {"error": "GRAPH_CLIENT_SECRET not set in Railway environment variables"}
    try:
        import requests as _req
    except ImportError:
        return {"error": "requests package not installed"}
    try:
        token = _get_graph_token()
    except Exception as e:
        return {"error": f"Auth failed: {e}"}

    hdrs = {"Authorization": f"Bearer {token}"}
    result = {"auth": "ok", "folders": [], "rmax_quotes_folder_id": None, "message_count": None}

    # List all top-level folders
    try:
        r = _req.get(f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders?$select=id,displayName&$top=100", headers=hdrs, timeout=20)
        result["folders_status"] = r.status_code
        result["folders_raw"] = r.text[:500]
        if r.status_code == 200:
            folders = r.json().get("value", [])
            result["folders"] = [f["displayName"] for f in folders]
            for f in folders:
                if (f.get("displayName") or "").strip().lower() == "rmax quotes":
                    result["rmax_quotes_folder_id"] = f["id"]
    except Exception as e:
        result["folder_list_error"] = str(e)

    # Also check Inbox children
    if not result["rmax_quotes_folder_id"]:
        try:
            r = _req.get(f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders/Inbox/childFolders?$select=id,displayName&$top=100", headers=hdrs, timeout=20)
            result["inbox_children_status"] = r.status_code
            result["inbox_children_raw"] = r.text[:500]
            if r.status_code == 200:
                children = r.json().get("value", [])
                result["inbox_children"] = [f["displayName"] for f in children]
                for f in children:
                    if (f.get("displayName") or "").strip().lower() == "rmax quotes":
                        result["rmax_quotes_folder_id"] = f["id"]
        except Exception as e:
            result["inbox_children_error"] = str(e)

    # If found, count messages
    if result["rmax_quotes_folder_id"]:
        fid = result["rmax_quotes_folder_id"]
        try:
            r = _req.get(f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders/{fid}/messages?$select=subject,receivedDateTime,from,toRecipients&$top=5", headers=hdrs, timeout=20)
            if r.status_code == 200:
                msgs = r.json().get("value", [])
                result["message_count"] = len(msgs)
                result["sample_emails"] = [{
                    "subject":   m.get("subject",""),
                    "date":      m.get("receivedDateTime","")[:10],
                    "from":      m.get("from",{}).get("emailAddress",{}).get("name",""),
                    "from_addr": m.get("from",{}).get("emailAddress",{}).get("address",""),
                    "to":        [r2.get("emailAddress",{}).get("name","") for r2 in m.get("toRecipients",[])],
                    "to_addr":   [r2.get("emailAddress",{}).get("address","") for r2 in m.get("toRecipients",[])],
                } for m in msgs[:3]]
        except Exception as e:
            result["message_fetch_error"] = str(e)

    return result

# ── Dashboard endpoint ────────────────────────────────────────────────────────
@app.get("/api/dashboard")
def dashboard():
    with get_db() as con:
        # KPI totals
        totals = con.execute("""
            SELECT
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total_amount,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won_amount,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal_amount,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost_amount
            FROM quotes WHERE (deleted IS NULL OR deleted=0)
        """).fetchone()

        # By location
        by_loc = con.execute("""
            SELECT location,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal
            FROM quotes
            WHERE (deleted IS NULL OR deleted=0) AND location IS NOT NULL
            GROUP BY location
            ORDER BY total DESC
        """).fetchall()

        # By region
        by_reg = con.execute("""
            SELECT COALESCE(NULLIF(TRIM(region),''),'(No Region)') as region,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal
            FROM quotes
            WHERE (deleted IS NULL OR deleted=0)
            GROUP BY region
            ORDER BY total DESC
        """).fetchall()

        # By month — parse dates in Python to handle M/D/YYYY, "Jul 13, 2026", etc.
        cutoff = datetime.now().replace(day=1) - timedelta(days=335)  # ~11 months ago
        raw_quotes = con.execute(
            "SELECT date_received, status, amount FROM quotes WHERE (deleted IS NULL OR deleted=0) AND date_received IS NOT NULL AND date_received != ''"
        ).fetchall()
        month_acc = defaultdict(lambda: dict(total_quotes=0,total=0.0,won=0.0,verbal=0.0,lost=0.0,won_count=0,verbal_count=0,lost_count=0))
        for row in raw_quotes:
            dt = parse_date(row["date_received"])
            if not dt or dt < cutoff:
                continue
            ym = dt.strftime('%Y-%m')
            amt = row["amount"] or 0
            st  = row["status"] or ''
            month_acc[ym]['total_quotes'] += 1
            month_acc[ym]['total']        += amt
            if st == 'Won':    month_acc[ym]['won']    += amt; month_acc[ym]['won_count']    += 1
            if st == 'Verbal': month_acc[ym]['verbal'] += amt; month_acc[ym]['verbal_count'] += 1
            if st == 'Lost':   month_acc[ym]['lost']   += amt; month_acc[ym]['lost_count']   += 1
        by_month = [{'month': k, **v} for k, v in sorted(month_acc.items())]

        # By close month — for 12-Month Rolling Projected Sales chart
        raw_close = con.execute(
            "SELECT close_date, status, amount FROM quotes "
            "WHERE (deleted IS NULL OR deleted=0) AND close_date IS NOT NULL AND close_date != ''"
        ).fetchall()
        close_acc = defaultdict(lambda: dict(total=0.0, won=0.0, verbal=0.0, open=0.0))
        for row in raw_close:
            dt = parse_date(row["close_date"])
            if not dt:
                continue
            ym  = dt.strftime('%Y-%m')
            amt = float(row["amount"] or 0)
            st  = (row["status"] or '').strip()
            if st in ('Lost', 'Duplicate'):
                continue
            close_acc[ym]['total'] += amt
            if st == 'Won':
                close_acc[ym]['won']    += amt
            elif st == 'Verbal':
                close_acc[ym]['verbal'] += amt
            else:
                close_acc[ym]['open']   += amt
        by_close_month = [{'month': k, 'total': round(v['total'], 2), 'won': round(v['won'], 2),
                           'verbal': round(v['verbal'], 2), 'open': round(v['open'], 2)}
                          for k, v in sorted(close_acc.items())]

        # Status breakdown
        by_status = con.execute("""
            SELECT status, COUNT(*) as count, COALESCE(SUM(amount),0) as amount
            FROM quotes
            WHERE (deleted IS NULL OR deleted=0) AND status IS NOT NULL
            GROUP BY status
            ORDER BY amount DESC
        """).fetchall()

        # By region and month — for regional monthly trends chart
        raw_rq = con.execute(
            "SELECT date_received, region, amount, status FROM quotes WHERE (deleted IS NULL OR deleted=0) AND date_received IS NOT NULL AND date_received != ''"
        ).fetchall()
        rm_acc = {}
        for row in raw_rq:
            dt = parse_date(row["date_received"])
            if not dt:
                continue
            ym     = dt.strftime('%Y-%m')
            region = (row["region"] or '').strip() or '(No Region)'
            key    = (region, ym)
            if key not in rm_acc:
                rm_acc[key] = {'region': region, 'month': ym, 'total': 0.0, 'count': 0, 'won': 0.0, 'verbal': 0.0}
            amt = float(row["amount"] or 0)
            rm_acc[key]['total'] += amt
            rm_acc[key]['count'] += 1
            st = (row["status"] or '').strip()
            if st == 'Won':
                rm_acc[key]['won'] += amt
            elif st == 'Verbal':
                rm_acc[key]['verbal'] += amt
        by_region_month = [{'region': k[0], 'month': k[1], 'total': round(v['total'], 2), 'count': v['count'],
                            'won': round(v['won'], 2), 'verbal': round(v['verbal'], 2)}
                           for k, v in sorted(rm_acc.items())]

        # Distinct filter options
        locations = [r[0] for r in con.execute(
            "SELECT DISTINCT location FROM quotes WHERE (deleted IS NULL OR deleted=0) AND location IS NOT NULL ORDER BY location"
        ).fetchall()]

        return {
            "totals": dict(totals),
            "by_location": [dict(r) for r in by_loc],
            "by_region": [dict(r) for r in by_reg],
            "by_region_month": by_region_month,
            "by_month": [dict(r) for r in by_month],
            "by_close_month": by_close_month,
            "by_status": [dict(r) for r in by_status],
            "locations": locations,
        }

# ── Overall Sales endpoint (reads live from Excel) ───────────────────────────
SALES_XLSX = os.path.join(BASE_DIR, "..", "Grow 2026 - SPECFORM Sales Plan.xlsx")

# ── Microsoft Graph / OneDrive config ─────────────────────────────────────────
# Set these as environment variables in Railway (never hardcode secrets in code)
GRAPH_TENANT_ID     = os.environ.get("GRAPH_TENANT_ID",     "10e8b460-4e9b-4c2b-b0e1-e102303252e1")
GRAPH_CLIENT_ID     = os.environ.get("GRAPH_CLIENT_ID",     "79b3406d-c46a-4128-912d-4b51c71f43a4")
GRAPH_CLIENT_SECRET = os.environ.get("GRAPH_CLIENT_SECRET", "")
GRAPH_USER          = os.environ.get("GRAPH_USER",          "mwalker@specformbc.com")
GRAPH_FILE_PATH     = os.environ.get("GRAPH_FILE_PATH",     "Documents/Specform Sales Partner/RMAX Weekly Quotes/Grow 2026 - SPECFORM Sales Plan.xlsx")

_token_cache: dict = {"token": None, "expires_at": 0.0}

def _get_graph_token() -> str:
    if _token_cache["token"] and time.time() < _token_cache["expires_at"] - 60:
        return _token_cache["token"]
    try:
        import msal
    except ImportError:
        raise HTTPException(500, "msal not installed — run: pip install msal")
    app = msal.ConfidentialClientApplication(
        GRAPH_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}",
        client_credential=GRAPH_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise HTTPException(500, f"Graph auth failed: {result.get('error_description', result)}")
    _token_cache["token"]      = result["access_token"]
    _token_cache["expires_at"] = time.time() + result.get("expires_in", 3600)
    return _token_cache["token"]

def _load_sales_workbook(data_only: bool = True, read_only: bool = False):
    """
    Return an openpyxl Workbook for the sales Excel file.
    Uses Microsoft Graph API when credentials are set; falls back to local file.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    if GRAPH_CLIENT_SECRET:
        try:
            import requests as _req
        except ImportError:
            raise HTTPException(500, "requests not installed — run: pip install requests")
        token = _get_graph_token()
        url   = (f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}"
                 f"/drive/root:/{GRAPH_FILE_PATH}:/content")
        resp  = _req.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
        if resp.status_code != 200:
            raise HTTPException(resp.status_code,
                f"OneDrive download failed ({resp.status_code}): {resp.text[:300]}")
        return openpyxl.load_workbook(io.BytesIO(resp.content),
                                      data_only=data_only, read_only=read_only)

    # ── Local file fallback ────────────────────────────────────────────────────
    xlsx_path = os.path.abspath(SALES_XLSX)
    if not os.path.exists(xlsx_path):
        alt = os.path.abspath(os.path.join(os.getcwd(), "..", "Grow 2026 - SPECFORM Sales Plan.xlsx"))
        if os.path.exists(alt):
            xlsx_path = alt
        else:
            raise HTTPException(404, f"Excel file not found at: {xlsx_path}")
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        shutil.copy2(xlsx_path, tmp_path)
        return openpyxl.load_workbook(tmp_path, data_only=data_only, read_only=read_only)
    except PermissionError:
        raise HTTPException(503, "Excel file is locked — close it in Excel and try again")
    except Exception as e:
        raise HTTPException(500, f"Failed to open Excel file: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try: os.unlink(tmp_path)
            except: pass

MONTH_SHEETS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
]
PRODUCTS = ["RMAX", "Hydrotech", "LAM", "Advanced Glassworks"]

def _norm_product(name: str) -> str:
    n = (name or "").strip().upper()
    if "HYDROTECH" in n or "HYDROTECH" in n: return "Hydrotech"
    if "LAM" in n:                            return "LAM"
    if "GLASS" in n:                          return "Advanced Glassworks"
    if "RMAX" in n:                           return "RMAX"
    return name.strip()

def _parse_sheet(ws):
    """Return dict with mtd_actual, mtd_plan, ytd_actual, annual_plan keyed by product."""
    # Discover product columns from the header row (col B = 'SALES')
    col_map = {}  # product_name → column index (0-based)
    for row in ws.iter_rows(max_row=4, values_only=True):
        if row[1] and str(row[1]).strip().upper() == "SALES":
            for i, cell in enumerate(row):
                if i > 1 and cell:
                    col_map[_norm_product(str(cell))] = i
            break

    result = {k: {p: 0 for p in PRODUCTS} for k in ("mtd_actual","mtd_plan","ytd_actual","annual_plan")}
    for row in ws.iter_rows(values_only=True):
        label = str(row[1] or "").strip()
        if not label: continue
        if "MTD" in label.upper() and "PROJECTED" in label.upper():
            key = "mtd_actual"
        elif "MTD" in label.upper() and "AVERAGE" in label.upper():
            key = "mtd_plan"
        elif "YTD" in label.upper() and "SALES" in label.upper():
            key = "ytd_actual"
        elif "PLAN" in label.upper() and "YEARLY" in label.upper():
            key = "annual_plan"
        else:
            continue
        for prod, idx in col_map.items():
            if prod in PRODUCTS:
                try:
                    val = row[idx]
                    result[key][prod] = float(val) if val is not None else 0
                except (TypeError, ValueError):
                    result[key][prod] = 0
    return result

@app.get("/api/sales")
def overall_sales():
    wb = _load_sales_workbook()
    try:
        months = []
        annual_plan = {p: 0 for p in PRODUCTS}

        for month in MONTH_SHEETS:
            if month not in wb.sheetnames:
                continue
            ws = wb[month]
            d  = _parse_sheet(ws)
            # Use the most recent annual plan found
            for p in PRODUCTS:
                if d["annual_plan"][p]:
                    annual_plan[p] = d["annual_plan"][p]

            mtd_total = sum(d["mtd_actual"].values())
            plan_total = sum(d["mtd_plan"].values())
            ytd_total  = sum(d["ytd_actual"].values())
            if mtd_total == 0 and plan_total == 0:
                continue  # skip months with no data at all

            months.append({
                "month":      month,
                "mtd_actual": d["mtd_actual"],
                "mtd_plan":   d["mtd_plan"],
                "ytd_actual": d["ytd_actual"],
                "mtd_total":  mtd_total,
                "plan_total": plan_total,
                "ytd_total":  ytd_total,
            })

        annual_total = sum(annual_plan.values())
        ytd_total_all = months[-1]["ytd_total"] if months else 0

        return {
            "months":       months,
            "annual_plan":  annual_plan,
            "annual_total": annual_total,
            "ytd_total":    ytd_total_all,
            "products":     PRODUCTS,
        }
    except Exception as e:
        raise HTTPException(500, f"Error parsing Excel data: {e}")

# ── Rep Sales endpoint ────────────────────────────────────────────────────────
REP_MONTH_SHEETS = [
    ("Jan",   "Jan per SBC Rep"),
    ("Feb",   "Feb per SBC Rep"),
    ("Mar",   "March per SBC Rep"),
    ("Apr",   "April per SBC Rep"),
    ("May",   "May per SBC Rep"),
    ("Jun",   "June per SBC Rep"),
    ("Jul",   "July per SBC Rep"),
    ("Aug",   "August per SBC Rep"),
    ("Sep",   "September per SBC Rep"),
    ("Oct",   "October per SBC Rep"),
    ("Nov",   "November per SBC Rep"),
    ("Dec",   "December per SBC Rep"),
]
REP_PRODUCTS = ["RMAX", "LAM", "American Hydrotech", "Advanced Glassworks"]

def _norm_rep_product(name: str) -> str:
    n = (name or "").strip().upper()
    if "RMAX"     in n:                          return "RMAX"
    if "LAM"      in n:                          return "LAM"
    if "HYDROTECH" in n or "AMERICAN" in n:      return "American Hydrotech"
    if "GLASS"    in n:                          return "Advanced Glassworks"
    return name.strip()

def _parse_rep_sheet(ws):
    """Parse a per-rep sheet. Returns mtd_plan, mtd_projected, and individual rep rows."""
    col_map = {}
    for row in ws.iter_rows(max_row=6, values_only=True):
        if row[1] and str(row[1]).strip().upper() == "SALES":
            for i, cell in enumerate(row):
                if i > 1 and cell:
                    col_map[_norm_rep_product(str(cell))] = i
            break
    if not col_map:
        return None

    def extract(row):
        out = {}
        for prod, idx in col_map.items():
            if prod in REP_PRODUCTS:
                try:
                    out[prod] = float(row[idx]) if idx < len(row) and row[idx] is not None else 0.0
                except (TypeError, ValueError):
                    out[prod] = 0.0
        return {p: out.get(p, 0.0) for p in REP_PRODUCTS}

    mtd_plan      = {p: 0.0 for p in REP_PRODUCTS}
    mtd_projected = {p: 0.0 for p in REP_PRODUCTS}
    reps = {}

    SKIP_PATTERNS = ["SALES", "GROW 2026", "PER SPECFORM", "PER SBC", "SALES PLAN", "YEARLY"]

    for row in ws.iter_rows(values_only=True):
        label = str(row[1] or "").strip()
        if not label:
            continue
        lu = label.upper()
        if any(p in lu for p in SKIP_PATTERNS):
            continue
        if "MTD" in lu and ("AVERAGE" in lu or "PLAN" in lu):
            mtd_plan = extract(row)
        elif "MTD" in lu and "PROJECTED" in lu:
            mtd_projected = extract(row)
        else:
            reps[label] = extract(row)

    return {"mtd_plan": mtd_plan, "mtd_projected": mtd_projected, "reps": reps}

@app.get("/api/rep-sales")
def rep_sales():
    wb = _load_sales_workbook()
    try:
        months = []
        for month_name, sheet_name in REP_MONTH_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            d = _parse_rep_sheet(wb[sheet_name])
            if d:
                months.append({"month": month_name, **d})
        return {"months": months, "products": REP_PRODUCTS}
    except Exception as e:
        raise HTTPException(500, f"Error parsing rep data: {e}")

@app.get("/api/sales/sheets")
def list_sheets():
    """Return all sheet names in the Excel file — used to identify rep tabs."""
    try:
        wb = _load_sales_workbook(read_only=True)
        return {"sheets": list(wb.sheetnames)}
    except Exception as e:
        raise HTTPException(500, str(e))

# ── Hydrotech Quotes ─────────────────────────────────────────────────────────

def init_hydrotech_db():
    with get_db() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS hydrotech_quotes (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            status             TEXT,
            date_received      TEXT,
            date_quoted        TEXT,
            sent_to            TEXT,
            subject            TEXT,
            job_name           TEXT,
            customer           TEXT,
            location           TEXT,
            product            TEXT,
            price              TEXT,
            quantities         TEXT,
            amount             REAL,
            close_date         TEXT,
            est_freight        TEXT,
            lead_time          TEXT,
            notes              TEXT,
            add_to_salesforce  INTEGER DEFAULT 0,
            completed          INTEGER DEFAULT 0,
            created_at         TEXT DEFAULT (datetime('now')),
            updated_at         TEXT DEFAULT (datetime('now'))
        )""")
        for col, definition in [('add_to_salesforce', 'INTEGER DEFAULT 0'),
                                 ('completed',         'INTEGER DEFAULT 0'),
                                 ('completed_date',    'TEXT'),
                                 ('region',            'TEXT'),
                                 ('deleted',           'INTEGER DEFAULT 0'),
                                 ('pdf_filename',      'TEXT'),
                                 ('company_id',        'INTEGER'),
                                 ('architect_id',      'INTEGER')]:
            try:
                con.execute(f"ALTER TABLE hydrotech_quotes ADD COLUMN {col} {definition}")
            except Exception:
                pass
        # Multi-PDF table
        con.execute("""
        CREATE TABLE IF NOT EXISTS hydrotech_quote_pdfs (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_id     INTEGER NOT NULL,
            pdf_filename TEXT NOT NULL,
            uploaded_at  TEXT DEFAULT (datetime('now'))
        )""")
        # Migrate any existing single pdf_filename values into the new table
        existing = con.execute(
            "SELECT id, pdf_filename FROM hydrotech_quotes WHERE pdf_filename IS NOT NULL AND pdf_filename != ''"
        ).fetchall()
        for row in existing:
            dup = con.execute(
                "SELECT id FROM hydrotech_quote_pdfs WHERE quote_id=? AND pdf_filename=?",
                (row["id"], row["pdf_filename"])
            ).fetchone()
            if not dup:
                con.execute(
                    "INSERT INTO hydrotech_quote_pdfs (quote_id, pdf_filename) VALUES (?,?)",
                    (row["id"], row["pdf_filename"])
                )
    # Ensure PDF storage directory exists
    os.makedirs(os.path.join(DATA_DIR, "hydrotech_pdfs"), exist_ok=True)

@app.get("/api/hydrotech-quotes")
def list_hydrotech_quotes(
    status:   Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    search:   Optional[str] = Query(None),
):
    with get_db() as con:
        sql = "SELECT * FROM hydrotech_quotes WHERE (deleted IS NULL OR deleted=0)"
        params = []
        if status and status != "All":
            sql += " AND status = ?"
            params.append(status)
        if location and location != "All":
            sql += " AND location = ?"
            params.append(location)
        if search:
            sql += " AND (subject LIKE ? OR sent_to LIKE ? OR job_name LIKE ? OR customer LIKE ? OR location LIKE ?)"
            s = f"%{search}%"
            params += [s, s, s, s, s]
        sql += " ORDER BY date_received DESC"
        rows = con.execute(sql, params).fetchall()
        quotes = [dict(r) for r in rows]
        if quotes:
            ids = [q["id"] for q in quotes]
            pdf_rows = con.execute(
                f"SELECT id, quote_id, pdf_filename FROM hydrotech_quote_pdfs WHERE quote_id IN ({','.join('?'*len(ids))}) ORDER BY uploaded_at",
                ids
            ).fetchall()
            pdf_map = {}
            for pr in pdf_rows:
                pdf_map.setdefault(pr["quote_id"], []).append({"id": pr["id"], "filename": pr["pdf_filename"]})
            for q in quotes:
                q["pdfs"] = pdf_map.get(q["id"], [])
        return quotes

@app.get("/api/hydrotech-quotes/{quote_id}")
def get_hydrotech_quote(quote_id: int):
    with get_db() as con:
        row = con.execute("SELECT * FROM hydrotech_quotes WHERE id=? AND (deleted IS NULL OR deleted=0)", (quote_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Hydrotech quote not found")
        q = dict(row)
        pdf_rows = con.execute(
            "SELECT id, pdf_filename FROM hydrotech_quote_pdfs WHERE quote_id=? ORDER BY uploaded_at", (quote_id,)
        ).fetchall()
        q["pdfs"] = [{"id": pr["id"], "filename": pr["pdf_filename"]} for pr in pdf_rows]
        return q

@app.post("/api/hydrotech-quotes", status_code=201)
def create_hydrotech_quote(q: QuoteIn):
    with get_db() as con:
        cur = con.execute("""
        INSERT INTO hydrotech_quotes (status,date_received,date_quoted,sent_to,subject,job_name,
            customer,location,product,price,quantities,amount,close_date,est_freight,lead_time,notes,
            region,add_to_salesforce,completed,company_id,architect_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (q.status,q.date_received,q.date_quoted,q.sent_to,q.subject,q.job_name,
              q.customer,q.location,q.product,q.price,q.quantities,q.amount,
              q.close_date,q.est_freight,q.lead_time,q.notes,
              q.region,q.add_to_salesforce,q.completed,q.company_id,q.architect_id))
        return {"id": cur.lastrowid}

@app.put("/api/hydrotech-quotes/{quote_id}")
def update_hydrotech_quote(quote_id: int, q: QuoteIn):
    with get_db() as con:
        existing = con.execute("SELECT completed, completed_date FROM hydrotech_quotes WHERE id=?", (quote_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Hydrotech quote not found")
        if q.completed and not existing['completed_date']:
            completed_date = date.today().strftime('%Y-%m-%d')
        elif not q.completed:
            completed_date = None
        else:
            completed_date = existing['completed_date']
        con.execute("""
        UPDATE hydrotech_quotes SET status=?,date_received=?,date_quoted=?,sent_to=?,subject=?,
            job_name=?,customer=?,location=?,product=?,price=?,quantities=?,amount=?,
            close_date=?,est_freight=?,lead_time=?,notes=?,
            region=?,add_to_salesforce=?,completed=?,completed_date=?,company_id=?,architect_id=?,updated_at=datetime('now')
        WHERE id=?
        """, (q.status,q.date_received,q.date_quoted,q.sent_to,q.subject,q.job_name,
              q.customer,q.location,q.product,q.price,q.quantities,q.amount,
              q.close_date,q.est_freight,q.lead_time,q.notes,
              q.region,q.add_to_salesforce,q.completed,completed_date,q.company_id,q.architect_id,quote_id))
        return {"ok": True}

@app.delete("/api/hydrotech-quotes/{quote_id}")
def delete_hydrotech_quote(quote_id: int):
    with get_db() as con:
        con.execute("UPDATE hydrotech_quotes SET deleted=1 WHERE id=?", (quote_id,))
        return {"ok": True}

@app.post("/api/hydrotech-quotes/{quote_id}/upload-pdf")
async def upload_hydrotech_pdf(quote_id: int, file: UploadFile):
    """Manually attach a PDF to an existing Hydrotech quote (supports multiple)."""
    os.makedirs(os.path.join(DATA_DIR, "hydrotech_pdfs"), exist_ok=True)
    safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in file.filename)
    pdf_filename = f"{quote_id}_{safe_name}"
    pdf_path = os.path.join(DATA_DIR, "hydrotech_pdfs", pdf_filename)
    contents = await file.read()
    with open(pdf_path, "wb") as f:
        f.write(contents)
    with get_db() as con:
        con.execute("INSERT INTO hydrotech_quote_pdfs (quote_id, pdf_filename) VALUES (?,?)", (quote_id, pdf_filename))
    return {"ok": True, "pdf_filename": pdf_filename}

@app.delete("/api/hydrotech-pdfs/{pdf_id}")
def delete_hydrotech_pdf(pdf_id: int):
    """Delete a single PDF attachment from a Hydrotech quote."""
    with get_db() as con:
        row = con.execute("SELECT pdf_filename FROM hydrotech_quote_pdfs WHERE id=?", (pdf_id,)).fetchone()
        if not row:
            raise HTTPException(404, "PDF not found")
        pdf_path = os.path.join(DATA_DIR, "hydrotech_pdfs", row["pdf_filename"])
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
        con.execute("DELETE FROM hydrotech_quote_pdfs WHERE id=?", (pdf_id,))
    return {"ok": True}

@app.get("/api/hydrotech-pdf/{filename}")
def serve_hydrotech_pdf(filename: str):
    """Serve a saved Hydrotech quote PDF attachment."""
    # Sanitize: no path traversal
    safe = os.path.basename(filename)
    pdf_path = os.path.join(DATA_DIR, "hydrotech_pdfs", safe)
    if not os.path.exists(pdf_path):
        raise HTTPException(404, "PDF not found")
    return FileResponse(pdf_path, media_type="application/pdf",
                        headers={"Content-Disposition": f'inline; filename="{safe}"'})

@app.get("/api/hydrotech-quotes-export")
def export_hydrotech_quotes():
    """Export all Hydrotech quotes to a formatted Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    with get_db() as con:
        rows = con.execute("""
            SELECT id, status, date_received, date_quoted, sent_to, subject,
                   job_name, customer, location, product, price, quantities,
                   amount, close_date, est_freight, lead_time, notes,
                   add_to_salesforce, completed
            FROM hydrotech_quotes ORDER BY date_received DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hydrotech Quotes"

    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='166534')   # dark green for Hydrotech
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_side  = Side(style='thin', color='BFBFBF')
    cell_border  = Border(left=border_side, right=border_side,
                          top=border_side, bottom=border_side)
    zebra_fill   = PatternFill('solid', fgColor='DCFCE7')
    STATUS_FILLS = {
        'Won':    PatternFill('solid', fgColor='D9EAD3'),
        'Lost':   PatternFill('solid', fgColor='F4CCCC'),
        'Verbal': PatternFill('solid', fgColor='FFF2CC'),
    }

    headers = [
        ('ID', 8), ('Status', 12), ('Date Received', 14), ('Date Quoted', 14),
        ('Sent To', 22), ('Subject', 30), ('Job Name', 28), ('Customer', 22),
        ('Location', 18), ('Product', 16), ('Price', 12), ('Quantities', 14),
        ('Amount', 14), ('Close Date', 14), ('Est. Freight', 14), ('Lead Time', 14),
        ('Notes', 35), ('Salesforce', 12), ('Completed', 12),
    ]
    ws.row_dimensions[1].height = 30
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = 'A2'

    for row_idx, q in enumerate(rows, start=2):
        status   = q['status'] or ''
        row_fill = STATUS_FILLS.get(status, (zebra_fill if row_idx % 2 == 0 else None))
        values = [
            q['id'], q['status'], q['date_received'], q['date_quoted'],
            q['sent_to'], q['subject'], q['job_name'], q['customer'],
            q['location'], q['product'], q['price'], q['quantities'],
            q['amount'], q['close_date'], q['est_freight'], q['lead_time'],
            q['notes'],
            'Yes' if q['add_to_salesforce'] else '',
            'Yes' if q['completed'] else '',
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = cell_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx in (6, 7, 17)))
            if row_fill:
                cell.fill = row_fill
            if col_idx == 13 and value is not None:
                cell.number_format = '$#,##0.00'
            if col_idx in (1, 2, 18, 19):
                cell.alignment = Alignment(horizontal='center', vertical='top')

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Hydrotech_Quotes_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.post("/api/hydrotech-sync")
def hydrotech_sync():
    """
    Pull new emails from the 'Hydrotech Quotes' Outlook folder via Graph API,
    parse each email body to pre-populate quote fields, and insert any
    that are not already in the database.
    """
    if not GRAPH_CLIENT_SECRET:
        return {"inserted": 0, "skipped": 0, "message": "Graph API not configured"}

    try:
        import requests as _req
    except ImportError:
        return {"inserted": 0, "skipped": 0, "message": "requests package not installed"}

    try:
        token = _get_graph_token()
    except Exception as e:
        return {"inserted": 0, "skipped": 0, "message": f"Auth failed: {e}"}

    hdrs = {"Authorization": f"Bearer {token}"}

    # ── Find "Hydrotech Quotes" folder ───────────────────────────────────────
    folder_id = None
    for list_url in [
        f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders?$select=id,displayName&$top=100",
        f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders/Inbox/childFolders?$select=id,displayName&$top=100",
    ]:
        try:
            r = _req.get(list_url, headers=hdrs, timeout=20)
            if r.status_code == 200:
                for f in r.json().get("value", []):
                    if (f.get("displayName") or "").strip().lower() == "hydrotech quotes":
                        folder_id = f["id"]
                        break
            if folder_id:
                break
        except Exception:
            pass

    if not folder_id:
        return {"inserted": 0, "skipped": 0,
                "message": "Could not find 'Hydrotech Quotes' folder in Outlook — check folder name"}

    # ── Fetch messages with body ──────────────────────────────────────────────
    msgs = []
    next_url = (f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}"
                f"/mailFolders/{folder_id}/messages"
                f"?$select=id,from,toRecipients,subject,receivedDateTime,body,hasAttachments&$top=50")
    while next_url and len(msgs) < 200:
        try:
            r = _req.get(next_url, headers=hdrs, timeout=30)
            if r.status_code != 200:
                return {"inserted": 0, "skipped": 0,
                        "message": f"Graph API error {r.status_code}: {r.text[:200]}"}
            data = r.json()
            msgs.extend(data.get("value", []))
            next_url = data.get("@odata.nextLink")
        except Exception as e:
            return {"inserted": 0, "skipped": 0, "message": f"Fetch error: {e}"}

    # ── Insert / update quotes ────────────────────────────────────────────────
    inserted = updated = skipped = 0
    with get_db() as con:
        for msg in msgs:
            # For Hydrotech, the customer is the *sender* of the incoming quote request
            from_addr = msg.get("from", {}).get("emailAddress", {})
            sent_to   = (from_addr.get("address") or from_addr.get("name") or "").strip()

            subject       = (msg.get("subject") or "").strip()
            received_raw  = msg.get("receivedDateTime", "")
            date_received = received_raw[:10] if received_raw else ""

            exists = con.execute(
                "SELECT id, sent_to FROM hydrotech_quotes WHERE subject=? AND date_received=?",
                (subject, date_received)
            ).fetchone()
            if exists:
                # Only fill in sent_to if blank — never overwrite a user-edited value
                if sent_to and not exists["sent_to"]:
                    con.execute("UPDATE hydrotech_quotes SET sent_to=? WHERE id=?",
                                (sent_to, exists["id"]))
                    updated += 1
                else:
                    skipped += 1
                continue

            body_content = msg.get("body", {}).get("content", "")
            content_type = msg.get("body", {}).get("contentType", "text")
            parsed = _parse_quote_email(subject, body_content, content_type)

            cur = con.execute("""
            INSERT INTO hydrotech_quotes
                (date_received, sent_to, subject, job_name, customer, location,
                 product, price, quantities, amount, est_freight, lead_time)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                date_received, sent_to, subject,
                parsed.get('job_name'), parsed.get('customer'), parsed.get('location'),
                parsed.get('product'), parsed.get('price'), parsed.get('quantities'),
                parsed.get('amount'), parsed.get('est_freight'), parsed.get('lead_time'),
            ))
            new_id = cur.lastrowid
            inserted += 1

            # ── Save PDF attachment if present ────────────────────────────────
            if msg.get("hasAttachments") and msg.get("id"):
                try:
                    att_url = (f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}"
                               f"/messages/{msg['id']}/attachments"
                               f"?$select=name,contentType,contentBytes&$top=20")
                    att_r = _req.get(att_url, headers=hdrs, timeout=30)
                    if att_r.status_code == 200:
                        for att in att_r.json().get("value", []):
                            ct = (att.get("contentType") or "").lower()
                            name = (att.get("name") or "attachment.pdf")
                            if "pdf" in ct or name.lower().endswith(".pdf"):
                                import base64 as _b64
                                pdf_bytes = _b64.b64decode(att.get("contentBytes", ""))
                                # Sanitize filename and prefix with quote id
                                safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in name)
                                pdf_filename = f"{new_id}_{safe_name}"
                                pdf_path = os.path.join(DATA_DIR, "hydrotech_pdfs", pdf_filename)
                                with open(pdf_path, "wb") as pf:
                                    pf.write(pdf_bytes)
                                con.execute("INSERT INTO hydrotech_quote_pdfs (quote_id, pdf_filename) VALUES (?,?)",
                                            (new_id, pdf_filename))
                                # continue loop to save ALL PDF attachments (not just first)
                except Exception:
                    pass  # don't fail the import if PDF save fails

    parts = []
    if inserted: parts.append(f"{inserted} new quote{'' if inserted==1 else 's'} imported")
    if updated:  parts.append(f"{updated} updated")
    msg_text = "✓ " + ", ".join(parts) if parts else "✓ Already up to date"
    return {"inserted": inserted, "skipped": skipped, "message": msg_text}

# ── Glassworks Quotes ────────────────────────────────────────────────────────

def init_glassworks_db():
    with get_db() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS glassworks_quotes (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            status             TEXT,
            date_received      TEXT,
            date_quoted        TEXT,
            sent_to            TEXT,
            subject            TEXT,
            job_name           TEXT,
            customer           TEXT,
            location           TEXT,
            product            TEXT,
            price              TEXT,
            quantities         TEXT,
            amount             REAL,
            close_date         TEXT,
            est_freight        TEXT,
            lead_time          TEXT,
            notes              TEXT,
            region             TEXT,
            add_to_salesforce  INTEGER DEFAULT 0,
            completed          INTEGER DEFAULT 0,
            deleted            INTEGER DEFAULT 0,
            created_at         TEXT DEFAULT (datetime('now')),
            updated_at         TEXT DEFAULT (datetime('now'))
        )""")
        for col, definition in [('add_to_salesforce', 'INTEGER DEFAULT 0'),
                                 ('completed',         'INTEGER DEFAULT 0'),
                                 ('completed_date',    'TEXT'),
                                 ('region',            'TEXT'),
                                 ('deleted',           'INTEGER DEFAULT 0'),
                                 ('company_id',        'INTEGER'),
                                 ('architect_id',      'INTEGER')]:
            try:
                con.execute(f"ALTER TABLE glassworks_quotes ADD COLUMN {col} {definition}")
            except Exception:
                pass
        con.execute("""
        CREATE TABLE IF NOT EXISTS glassworks_quote_files (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_id      INTEGER NOT NULL,
            file_filename TEXT NOT NULL,
            uploaded_at   TEXT DEFAULT (datetime('now'))
        )""")

@app.get("/api/glassworks-quotes")
def list_glassworks_quotes(
    status:   Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    search:   Optional[str] = Query(None),
):
    with get_db() as con:
        sql = "SELECT * FROM glassworks_quotes WHERE (deleted IS NULL OR deleted=0)"
        params = []
        if status and status != "All":
            sql += " AND status = ?"
            params.append(status)
        if location and location != "All":
            sql += " AND location = ?"
            params.append(location)
        if search:
            sql += " AND (subject LIKE ? OR sent_to LIKE ? OR job_name LIKE ? OR customer LIKE ? OR location LIKE ?)"
            s = f"%{search}%"
            params += [s, s, s, s, s]
        sql += " ORDER BY date_received DESC"
        rows = con.execute(sql, params).fetchall()
        quotes = [dict(r) for r in rows]
        if quotes:
            ids = [q['id'] for q in quotes]
            file_rows = con.execute(
                f"SELECT id, quote_id, file_filename FROM glassworks_quote_files WHERE quote_id IN ({','.join('?'*len(ids))})",
                ids
            ).fetchall()
            file_map = {}
            for fr in file_rows:
                file_map.setdefault(fr['quote_id'], []).append({'id': fr['id'], 'filename': fr['file_filename']})
            for q in quotes:
                q['files'] = file_map.get(q['id'], [])
        return quotes

@app.get("/api/glassworks-quotes/{quote_id}")
def get_glassworks_quote(quote_id: int):
    with get_db() as con:
        row = con.execute("SELECT * FROM glassworks_quotes WHERE id=? AND (deleted IS NULL OR deleted=0)", (quote_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Glassworks quote not found")
        return dict(row)

@app.post("/api/glassworks-quotes", status_code=201)
def create_glassworks_quote(q: QuoteIn):
    with get_db() as con:
        cur = con.execute("""
        INSERT INTO glassworks_quotes (status,date_received,date_quoted,sent_to,subject,job_name,
            customer,location,product,price,quantities,amount,close_date,est_freight,lead_time,notes,
            region,add_to_salesforce,completed,company_id,architect_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (q.status,normalize_date(q.date_received),normalize_date(q.date_quoted),
              q.sent_to,q.subject,q.job_name,
              q.customer,q.location,q.product,q.price,q.quantities,q.amount,
              normalize_date(q.close_date),q.est_freight,q.lead_time,q.notes,
              q.region,q.add_to_salesforce,q.completed,q.company_id,q.architect_id))
        return {"id": cur.lastrowid}

@app.put("/api/glassworks-quotes/{quote_id}")
def update_glassworks_quote(quote_id: int, q: QuoteIn):
    with get_db() as con:
        existing = con.execute("SELECT completed, completed_date FROM glassworks_quotes WHERE id=?", (quote_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Glassworks quote not found")
        if q.completed and not existing['completed_date']:
            completed_date = date.today().strftime('%Y-%m-%d')
        elif not q.completed:
            completed_date = None
        else:
            completed_date = existing['completed_date']
        con.execute("""
        UPDATE glassworks_quotes SET status=?,date_received=?,date_quoted=?,sent_to=?,subject=?,
            job_name=?,customer=?,location=?,product=?,price=?,quantities=?,amount=?,
            close_date=?,est_freight=?,lead_time=?,notes=?,
            region=?,add_to_salesforce=?,completed=?,completed_date=?,company_id=?,architect_id=?,updated_at=datetime('now')
        WHERE id=?
        """, (q.status,normalize_date(q.date_received),normalize_date(q.date_quoted),
              q.sent_to,q.subject,q.job_name,
              q.customer,q.location,q.product,q.price,q.quantities,q.amount,
              normalize_date(q.close_date),q.est_freight,q.lead_time,q.notes,
              q.region,q.add_to_salesforce,q.completed,completed_date,q.company_id,q.architect_id,quote_id))
        return {"ok": True}

@app.delete("/api/glassworks-quotes/{quote_id}")
def delete_glassworks_quote(quote_id: int):
    with get_db() as con:
        con.execute("UPDATE glassworks_quotes SET deleted=1 WHERE id=?", (quote_id,))
        return {"ok": True}

@app.get("/api/glassworks-quotes-export")
def export_glassworks_quotes():
    """Export all Glassworks quotes to a formatted Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    import io

    with get_db() as con:
        rows = con.execute("""
            SELECT id, status, date_received, date_quoted, sent_to, subject,
                   job_name, customer, location, product, price, quantities,
                   amount, close_date, est_freight, lead_time, notes,
                   add_to_salesforce, completed
            FROM glassworks_quotes WHERE (deleted IS NULL OR deleted=0) ORDER BY date_received DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Glassworks Quotes"

    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='0e7490')   # teal for Glassworks
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_side  = Side(style='thin', color='BFBFBF')
    cell_border  = Border(left=border_side, right=border_side,
                          top=border_side, bottom=border_side)
    zebra_fill   = PatternFill('solid', fgColor='CFFAFE')
    STATUS_FILLS = {
        'Won':    PatternFill('solid', fgColor='D9EAD3'),
        'Lost':   PatternFill('solid', fgColor='F4CCCC'),
        'Verbal': PatternFill('solid', fgColor='FFF2CC'),
    }

    headers = [
        ('ID', 8), ('Status', 12), ('Date Received', 14), ('Date Quoted', 14),
        ('Sent To', 22), ('Subject', 30), ('Job Name', 28), ('Customer', 22),
        ('Location', 18), ('Product', 16), ('Price', 12), ('Quantities', 14),
        ('Amount', 14), ('Close Date', 14), ('Est. Freight', 14), ('Lead Time', 14),
        ('Notes', 35), ('Salesforce', 12), ('Completed', 12),
    ]
    ws.row_dimensions[1].height = 30
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = 'A2'

    for row_idx, q in enumerate(rows, start=2):
        status   = q['status'] or ''
        row_fill = STATUS_FILLS.get(status, (zebra_fill if row_idx % 2 == 0 else None))
        values = [
            q['id'], q['status'], q['date_received'], q['date_quoted'],
            q['sent_to'], q['subject'], q['job_name'], q['customer'],
            q['location'], q['product'], q['price'], q['quantities'],
            q['amount'], q['close_date'], q['est_freight'], q['lead_time'],
            q['notes'],
            'Yes' if q['add_to_salesforce'] else '',
            'Yes' if q['completed'] else '',
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = cell_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx in (6, 7, 17)))
            if row_fill:
                cell.fill = row_fill
            if col_idx == 13 and value is not None:
                cell.number_format = '$#,##0.00'
            if col_idx in (1, 2, 18, 19):
                cell.alignment = Alignment(horizontal='center', vertical='top')

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Glassworks_Quotes_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.post("/api/glassworks-quotes/{quote_id}/upload-file")
async def upload_glassworks_file(quote_id: int, file: UploadFile):
    """Attach a file to an existing Glassworks quote."""
    import mimetypes
    os.makedirs(os.path.join(DATA_DIR, 'glassworks_files'), exist_ok=True)
    safe_name = ''.join(c if c.isalnum() or c in '._- ' else '_' for c in file.filename)
    file_filename = f'{quote_id}_{safe_name}'
    file_path = os.path.join(DATA_DIR, 'glassworks_files', file_filename)
    contents = await file.read()
    with open(file_path, 'wb') as f:
        f.write(contents)
    with get_db() as con:
        con.execute('INSERT INTO glassworks_quote_files (quote_id, file_filename) VALUES (?,?)', (quote_id, file_filename))
    return {'ok': True, 'file_filename': file_filename}

@app.delete("/api/glassworks-files/{file_id}")
def delete_glassworks_file(file_id: int):
    """Delete a file attachment from a Glassworks quote."""
    with get_db() as con:
        row = con.execute('SELECT file_filename FROM glassworks_quote_files WHERE id=?', (file_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail='File not found')
        file_path = os.path.join(DATA_DIR, 'glassworks_files', row['file_filename'])
        if os.path.exists(file_path):
            os.remove(file_path)
        con.execute('DELETE FROM glassworks_quote_files WHERE id=?', (file_id,))
    return {'ok': True}

@app.get("/api/glassworks-file/{filename}")
def serve_glassworks_file(filename: str):
    """Serve a saved Glassworks quote file attachment."""
    import mimetypes
    safe = os.path.basename(filename)
    file_path = os.path.join(DATA_DIR, 'glassworks_files', safe)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail='File not found')
    mt = mimetypes.guess_type(safe)[0] or 'application/octet-stream'
    display_name = safe.split('_', 1)[-1] if '_' in safe else safe
    return FileResponse(file_path, media_type=mt, filename=display_name)

@app.post("/api/glassworks-sync")
def glassworks_sync():
    """Pull new emails from 'Glassworks Quotes' Outlook folder via Graph API."""
    if not GRAPH_CLIENT_SECRET:
        return {"inserted": 0, "skipped": 0, "message": "Graph API not configured"}

    try:
        import requests as _req
    except ImportError:
        return {"inserted": 0, "skipped": 0, "message": "requests package not installed"}

    try:
        token = _get_graph_token()
    except Exception as e:
        return {"inserted": 0, "skipped": 0, "message": f"Auth failed: {e}"}

    hdrs = {"Authorization": f"Bearer {token}"}

    folder_id = None
    for list_url in [
        f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders?$select=id,displayName&$top=100",
        f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders/Inbox/childFolders?$select=id,displayName&$top=100",
    ]:
        try:
            r = _req.get(list_url, headers=hdrs, timeout=20)
            if r.status_code == 200:
                for f in r.json().get("value", []):
                    if (f.get("displayName") or "").strip().lower() == "glassworks quotes":
                        folder_id = f["id"]
                        break
            if folder_id:
                break
        except Exception:
            pass

    if not folder_id:
        return {"inserted": 0, "skipped": 0,
                "message": "Could not find 'Glassworks Quotes' folder in Outlook — check folder name"}

    msgs = []
    next_url = (f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}"
                f"/mailFolders/{folder_id}/messages"
                f"?$select=from,toRecipients,subject,receivedDateTime,body&$top=50")
    while next_url and len(msgs) < 200:
        try:
            r = _req.get(next_url, headers=hdrs, timeout=30)
            if r.status_code != 200:
                return {"inserted": 0, "skipped": 0,
                        "message": f"Graph API error {r.status_code}: {r.text[:200]}"}
            data = r.json()
            msgs.extend(data.get("value", []))
            next_url = data.get("@odata.nextLink")
        except Exception as e:
            return {"inserted": 0, "skipped": 0, "message": f"Fetch error: {e}"}

    inserted = updated = skipped = 0
    with get_db() as con:
        for msg in msgs:
            to_recipients = msg.get("toRecipients", [])
            if to_recipients:
                first_to = to_recipients[0].get("emailAddress", {})
                sent_to  = (first_to.get("name") or first_to.get("address") or "").strip()
            else:
                sent_to  = ""

            subject       = (msg.get("subject") or "").strip()
            received_raw  = msg.get("receivedDateTime", "")
            date_received = received_raw[:10] if received_raw else ""

            exists = con.execute(
                "SELECT id, sent_to FROM glassworks_quotes WHERE subject=? AND date_received=?",
                (subject, date_received)
            ).fetchone()
            if exists:
                # Only fill in sent_to if blank — never overwrite a user-edited value
                if sent_to and not exists["sent_to"]:
                    con.execute("UPDATE glassworks_quotes SET sent_to=? WHERE id=?",
                                (sent_to, exists["id"]))
                    updated += 1
                else:
                    skipped += 1
                continue

            body_content = msg.get("body", {}).get("content", "")
            content_type = msg.get("body", {}).get("contentType", "text")
            parsed = _parse_quote_email(subject, body_content, content_type)

            con.execute("""
            INSERT INTO glassworks_quotes
                (date_received, sent_to, subject, job_name, customer, location,
                 product, price, quantities, amount, est_freight, lead_time)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                date_received, sent_to, subject,
                parsed.get('job_name'), parsed.get('customer'), parsed.get('location'),
                parsed.get('product'), parsed.get('price'), parsed.get('quantities'),
                parsed.get('amount'), parsed.get('est_freight'), parsed.get('lead_time'),
            ))
            inserted += 1

    parts = []
    if inserted: parts.append(f"{inserted} new quote{'' if inserted==1 else 's'} imported")
    if updated:  parts.append(f"{updated} updated")
    msg_text = "✓ " + ", ".join(parts) if parts else "✓ Already up to date"
    return {"inserted": inserted, "skipped": skipped, "message": msg_text}

@app.get("/api/hydrotech-dashboard")
def hydrotech_dashboard():
    with get_db() as con:
        totals = con.execute("""
            SELECT
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total_amount,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won_amount,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal_amount,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost_amount
            FROM hydrotech_quotes
        """).fetchone()
        by_loc = con.execute("""
            SELECT location,
                COUNT(*) as count,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal
            FROM hydrotech_quotes WHERE location IS NOT NULL
            GROUP BY location ORDER BY total DESC
        """).fetchall()
        by_month = con.execute("""
            SELECT substr(date_received,1,7) as month,
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count
            FROM hydrotech_quotes
            WHERE date_received IS NOT NULL AND date_received != ''
              AND substr(date_received,1,7) >= substr(date('now','-11 months'),1,7)
            GROUP BY month ORDER BY month
        """).fetchall()
        by_status = con.execute("""
            SELECT status, COUNT(*) as count, COALESCE(SUM(amount),0) as amount
            FROM hydrotech_quotes WHERE status IS NOT NULL
            GROUP BY status ORDER BY amount DESC
        """).fetchall()
        locations = [r[0] for r in con.execute(
            "SELECT DISTINCT location FROM hydrotech_quotes WHERE location IS NOT NULL ORDER BY location"
        ).fetchall()]
        raw_close_ht = con.execute(
            "SELECT close_date, status, amount FROM hydrotech_quotes "
            "WHERE (deleted IS NULL OR deleted=0) AND close_date IS NOT NULL AND close_date != ''"
        ).fetchall()
        from collections import defaultdict as _dd2
        close_acc_ht = _dd2(lambda: dict(total=0.0, won=0.0, verbal=0.0, open=0.0))
        for row in raw_close_ht:
            dt = parse_date(row["close_date"])
            if not dt: continue
            ym = dt.strftime('%Y-%m')
            amt = float(row["amount"] or 0)
            st = (row["status"] or '').strip()
            if st in ('Lost', 'Duplicate'): continue
            close_acc_ht[ym]['total'] += amt
            if st == 'Won':     close_acc_ht[ym]['won']    += amt
            elif st == 'Verbal': close_acc_ht[ym]['verbal'] += amt
            else:                close_acc_ht[ym]['open']   += amt
        by_close_month_ht = [{'month': k, 'total': round(v['total'], 2), 'won': round(v['won'], 2),
                               'verbal': round(v['verbal'], 2), 'open': round(v['open'], 2)}
                              for k, v in sorted(close_acc_ht.items())]
        return {
            "totals": dict(totals),
            "by_location": [dict(r) for r in by_loc],
            "by_month": [dict(r) for r in by_month],
            "by_status": [dict(r) for r in by_status],
            "locations": locations,
            "by_close_month": by_close_month_ht,
        }

@app.get("/api/glassworks-dashboard")
def glassworks_dashboard():
    with get_db() as con:
        totals = con.execute("""
            SELECT
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total_amount,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won_amount,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal_amount,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost_amount
            FROM glassworks_quotes WHERE (deleted IS NULL OR deleted=0)
        """).fetchone()
        by_loc = con.execute("""
            SELECT location,
                COUNT(*) as count,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal
            FROM glassworks_quotes WHERE (deleted IS NULL OR deleted=0) AND location IS NOT NULL
            GROUP BY location ORDER BY total DESC
        """).fetchall()
        by_month = con.execute("""
            SELECT substr(date_received,1,7) as month,
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count
            FROM glassworks_quotes
            WHERE (deleted IS NULL OR deleted=0) AND date_received IS NOT NULL AND date_received != ''
              AND substr(date_received,1,7) >= substr(date('now','-11 months'),1,7)
            GROUP BY month ORDER BY month
        """).fetchall()
        by_status = con.execute("""
            SELECT status, COUNT(*) as count, COALESCE(SUM(amount),0) as amount
            FROM glassworks_quotes WHERE (deleted IS NULL OR deleted=0) AND status IS NOT NULL
            GROUP BY status ORDER BY amount DESC
        """).fetchall()
        raw_close_gw = con.execute(
            "SELECT close_date, status, amount FROM glassworks_quotes "
            "WHERE (deleted IS NULL OR deleted=0) AND close_date IS NOT NULL AND close_date != ''"
        ).fetchall()
        from collections import defaultdict as _dd3
        close_acc_gw = _dd3(lambda: dict(total=0.0, won=0.0, verbal=0.0, open=0.0))
        for row in raw_close_gw:
            dt = parse_date(row["close_date"])
            if not dt: continue
            ym = dt.strftime('%Y-%m')
            amt = float(row["amount"] or 0)
            st = (row["status"] or '').strip()
            if st in ('Lost', 'Duplicate'): continue
            close_acc_gw[ym]['total'] += amt
            if st == 'Won':      close_acc_gw[ym]['won']    += amt
            elif st == 'Verbal': close_acc_gw[ym]['verbal'] += amt
            else:                close_acc_gw[ym]['open']   += amt
        by_close_month_gw = [{'month': k, 'total': round(v['total'], 2), 'won': round(v['won'], 2),
                               'verbal': round(v['verbal'], 2), 'open': round(v['open'], 2)}
                              for k, v in sorted(close_acc_gw.items())]
        return {
            "totals": dict(totals),
            "by_location": [dict(r) for r in by_loc],
            "by_month": [dict(r) for r in by_month],
            "by_status": [dict(r) for r in by_status],
            "by_close_month": by_close_month_gw,
        }


def init_lam_db():
    with get_db() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS lam_quotes (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            status             TEXT,
            date_received      TEXT,
            date_quoted        TEXT,
            sent_to            TEXT,
            subject            TEXT,
            job_name           TEXT,
            customer           TEXT,
            location           TEXT,
            product            TEXT,
            price              TEXT,
            quantities         TEXT,
            amount             REAL,
            close_date         TEXT,
            est_freight        TEXT,
            lead_time          TEXT,
            notes              TEXT,
            region             TEXT,
            add_to_salesforce  INTEGER DEFAULT 0,
            completed          INTEGER DEFAULT 0,
            deleted            INTEGER DEFAULT 0,
            created_at         TEXT DEFAULT (datetime('now')),
            updated_at         TEXT DEFAULT (datetime('now'))
        )""")
        for col, definition in [('add_to_salesforce', 'INTEGER DEFAULT 0'),
                                 ('completed',         'INTEGER DEFAULT 0'),
                                 ('completed_date',    'TEXT'),
                                 ('region',            'TEXT'),
                                 ('deleted',           'INTEGER DEFAULT 0'),
                                 ('company_id',        'INTEGER'),
                                 ('architect_id',      'INTEGER')]:
            try:
                con.execute(f"ALTER TABLE lam_quotes ADD COLUMN {col} {definition}")
            except Exception:
                pass
        con.execute("""
        CREATE TABLE IF NOT EXISTS lam_quote_files (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_id      INTEGER NOT NULL,
            file_filename TEXT NOT NULL,
            uploaded_at   TEXT DEFAULT (datetime('now'))
        )""")

@app.get("/api/lam-quotes")
def list_lam_quotes(
    status:   Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    search:   Optional[str] = Query(None),
):
    with get_db() as con:
        sql = "SELECT * FROM lam_quotes WHERE (deleted IS NULL OR deleted=0)"
        params = []
        if status and status != "All":
            sql += " AND status = ?"
            params.append(status)
        if location and location != "All":
            sql += " AND location = ?"
            params.append(location)
        if search:
            sql += " AND (subject LIKE ? OR sent_to LIKE ? OR job_name LIKE ? OR customer LIKE ? OR location LIKE ?)"
            s = f"%{search}%"
            params += [s, s, s, s, s]
        sql += " ORDER BY date_received DESC"
        rows = con.execute(sql, params).fetchall()
        quotes = [dict(r) for r in rows]
        if quotes:
            ids = [q['id'] for q in quotes]
            file_rows = con.execute(
                f"SELECT id, quote_id, file_filename FROM lam_quote_files WHERE quote_id IN ({','.join('?'*len(ids))})",
                ids
            ).fetchall()
            file_map = {}
            for fr in file_rows:
                file_map.setdefault(fr['quote_id'], []).append({'id': fr['id'], 'filename': fr['file_filename']})
            for q in quotes:
                q['files'] = file_map.get(q['id'], [])
        return quotes

@app.get("/api/lam-quotes/{quote_id}")
def get_lam_quote(quote_id: int):
    with get_db() as con:
        row = con.execute("SELECT * FROM lam_quotes WHERE id=? AND (deleted IS NULL OR deleted=0)", (quote_id,)).fetchone()
        if not row:
            raise HTTPException(404, "LAM quote not found")
        return dict(row)

@app.post("/api/lam-quotes", status_code=201)
def create_lam_quote(q: QuoteIn):
    with get_db() as con:
        cur = con.execute("""
        INSERT INTO lam_quotes (status,date_received,date_quoted,sent_to,subject,job_name,
            customer,location,product,price,quantities,amount,close_date,est_freight,lead_time,notes,
            region,add_to_salesforce,completed,company_id,architect_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (q.status,normalize_date(q.date_received),normalize_date(q.date_quoted),
              q.sent_to,q.subject,q.job_name,
              q.customer,q.location,q.product,q.price,q.quantities,q.amount,
              normalize_date(q.close_date),q.est_freight,q.lead_time,q.notes,
              q.region,q.add_to_salesforce,q.completed,q.company_id,q.architect_id))
        return {"id": cur.lastrowid}

@app.put("/api/lam-quotes/{quote_id}")
def update_lam_quote(quote_id: int, q: QuoteIn):
    with get_db() as con:
        existing = con.execute("SELECT completed, completed_date FROM lam_quotes WHERE id=?", (quote_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "LAM quote not found")
        if q.completed and not existing['completed_date']:
            completed_date = date.today().strftime('%Y-%m-%d')
        elif not q.completed:
            completed_date = None
        else:
            completed_date = existing['completed_date']
        con.execute("""
        UPDATE lam_quotes SET status=?,date_received=?,date_quoted=?,sent_to=?,subject=?,
            job_name=?,customer=?,location=?,product=?,price=?,quantities=?,amount=?,
            close_date=?,est_freight=?,lead_time=?,notes=?,
            region=?,add_to_salesforce=?,completed=?,completed_date=?,company_id=?,architect_id=?,updated_at=datetime('now')
        WHERE id=?
        """, (q.status,normalize_date(q.date_received),normalize_date(q.date_quoted),
              q.sent_to,q.subject,q.job_name,
              q.customer,q.location,q.product,q.price,q.quantities,q.amount,
              normalize_date(q.close_date),q.est_freight,q.lead_time,q.notes,
              q.region,q.add_to_salesforce,q.completed,completed_date,q.company_id,q.architect_id,quote_id))
        return {"ok": True}

@app.delete("/api/lam-quotes/{quote_id}")
def delete_lam_quote(quote_id: int):
    with get_db() as con:
        con.execute("UPDATE lam_quotes SET deleted=1 WHERE id=?", (quote_id,))
        return {"ok": True}

@app.get("/api/lam-quotes-export")
def export_lam_quotes():
    """Export all LAM quotes to a formatted Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    import io

    with get_db() as con:
        rows = con.execute("""
            SELECT id, status, date_received, date_quoted, sent_to, subject,
                   job_name, customer, location, product, price, quantities,
                   amount, close_date, est_freight, lead_time, notes,
                   add_to_salesforce, completed
            FROM lam_quotes WHERE (deleted IS NULL OR deleted=0) ORDER BY date_received DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LAM Quotes"

    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='0e7490')   # teal for LAM
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_side  = Side(style='thin', color='BFBFBF')
    cell_border  = Border(left=border_side, right=border_side,
                          top=border_side, bottom=border_side)
    zebra_fill   = PatternFill('solid', fgColor='CFFAFE')
    STATUS_FILLS = {
        'Won':    PatternFill('solid', fgColor='D9EAD3'),
        'Lost':   PatternFill('solid', fgColor='F4CCCC'),
        'Verbal': PatternFill('solid', fgColor='FFF2CC'),
    }

    headers = [
        ('ID', 8), ('Status', 12), ('Date Received', 14), ('Date Quoted', 14),
        ('Sent To', 22), ('Subject', 30), ('Job Name', 28), ('Customer', 22),
        ('Location', 18), ('Product', 16), ('Price', 12), ('Quantities', 14),
        ('Amount', 14), ('Close Date', 14), ('Est. Freight', 14), ('Lead Time', 14),
        ('Notes', 35), ('Salesforce', 12), ('Completed', 12),
    ]
    ws.row_dimensions[1].height = 30
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = 'A2'

    for row_idx, q in enumerate(rows, start=2):
        status   = q['status'] or ''
        row_fill = STATUS_FILLS.get(status, (zebra_fill if row_idx % 2 == 0 else None))
        values = [
            q['id'], q['status'], q['date_received'], q['date_quoted'],
            q['sent_to'], q['subject'], q['job_name'], q['customer'],
            q['location'], q['product'], q['price'], q['quantities'],
            q['amount'], q['close_date'], q['est_freight'], q['lead_time'],
            q['notes'],
            'Yes' if q['add_to_salesforce'] else '',
            'Yes' if q['completed'] else '',
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = cell_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx in (6, 7, 17)))
            if row_fill:
                cell.fill = row_fill
            if col_idx == 13 and value is not None:
                cell.number_format = '$#,##0.00'
            if col_idx in (1, 2, 18, 19):
                cell.alignment = Alignment(horizontal='center', vertical='top')

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"LAM_Quotes_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.post("/api/lam-quotes/{quote_id}/upload-file")
async def upload_lam_file(quote_id: int, file: UploadFile):
    """Attach a file to an existing LAM quote."""
    import mimetypes
    os.makedirs(os.path.join(DATA_DIR, 'lam_files'), exist_ok=True)
    safe_name = ''.join(c if c.isalnum() or c in '._- ' else '_' for c in file.filename)
    file_filename = f'{quote_id}_{safe_name}'
    file_path = os.path.join(DATA_DIR, 'lam_files', file_filename)
    contents = await file.read()
    with open(file_path, 'wb') as f:
        f.write(contents)
    with get_db() as con:
        con.execute('INSERT INTO lam_quote_files (quote_id, file_filename) VALUES (?,?)', (quote_id, file_filename))
    return {'ok': True, 'file_filename': file_filename}

@app.delete("/api/lam-files/{file_id}")
def delete_lam_file(file_id: int):
    """Delete a file attachment from a LAM quote."""
    with get_db() as con:
        row = con.execute('SELECT file_filename FROM lam_quote_files WHERE id=?', (file_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail='File not found')
        file_path = os.path.join(DATA_DIR, 'lam_files', row['file_filename'])
        if os.path.exists(file_path):
            os.remove(file_path)
        con.execute('DELETE FROM lam_quote_files WHERE id=?', (file_id,))
    return {'ok': True}

@app.get("/api/lam-file/{filename}")
def serve_lam_file(filename: str):
    """Serve a saved LAM quote file attachment."""
    import mimetypes
    safe = os.path.basename(filename)
    file_path = os.path.join(DATA_DIR, 'lam_files', safe)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail='File not found')
    mt = mimetypes.guess_type(safe)[0] or 'application/octet-stream'
    display_name = safe.split('_', 1)[-1] if '_' in safe else safe
    return FileResponse(file_path, media_type=mt, filename=display_name)

@app.post("/api/lam-sync")
def lam_sync():
    """Pull new emails from 'LAM Quotes' Outlook folder via Graph API."""
    if not GRAPH_CLIENT_SECRET:
        return {"inserted": 0, "skipped": 0, "message": "Graph API not configured"}

    try:
        import requests as _req
    except ImportError:
        return {"inserted": 0, "skipped": 0, "message": "requests package not installed"}

    try:
        token = _get_graph_token()
    except Exception as e:
        return {"inserted": 0, "skipped": 0, "message": f"Auth failed: {e}"}

    hdrs = {"Authorization": f"Bearer {token}"}

    folder_id = None
    for list_url in [
        f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders?$select=id,displayName&$top=100",
        f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders/Inbox/childFolders?$select=id,displayName&$top=100",
    ]:
        try:
            r = _req.get(list_url, headers=hdrs, timeout=20)
            if r.status_code == 200:
                for f in r.json().get("value", []):
                    if (f.get("displayName") or "").strip().lower() == "lam quotes":
                        folder_id = f["id"]
                        break
            if folder_id:
                break
        except Exception:
            pass

    if not folder_id:
        return {"inserted": 0, "skipped": 0,
                "message": "Could not find 'LAM Quotes' folder in Outlook — check folder name"}

    msgs = []
    next_url = (f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}"
                f"/mailFolders/{folder_id}/messages"
                f"?$select=from,toRecipients,subject,receivedDateTime,body&$top=50")
    while next_url and len(msgs) < 200:
        try:
            r = _req.get(next_url, headers=hdrs, timeout=30)
            if r.status_code != 200:
                return {"inserted": 0, "skipped": 0,
                        "message": f"Graph API error {r.status_code}: {r.text[:200]}"}
            data = r.json()
            msgs.extend(data.get("value", []))
            next_url = data.get("@odata.nextLink")
        except Exception as e:
            return {"inserted": 0, "skipped": 0, "message": f"Fetch error: {e}"}

    inserted = updated = skipped = 0
    with get_db() as con:
        for msg in msgs:
            to_recipients = msg.get("toRecipients", [])
            if to_recipients:
                first_to = to_recipients[0].get("emailAddress", {})
                sent_to  = (first_to.get("name") or first_to.get("address") or "").strip()
            else:
                sent_to  = ""

            subject       = (msg.get("subject") or "").strip()
            received_raw  = msg.get("receivedDateTime", "")
            date_received = received_raw[:10] if received_raw else ""

            exists = con.execute(
                "SELECT id, sent_to FROM lam_quotes WHERE subject=? AND date_received=?",
                (subject, date_received)
            ).fetchone()
            if exists:
                # Only fill in sent_to if blank — never overwrite a user-edited value
                if sent_to and not exists["sent_to"]:
                    con.execute("UPDATE lam_quotes SET sent_to=? WHERE id=?",
                                (sent_to, exists["id"]))
                    updated += 1
                else:
                    skipped += 1
                continue

            body_content = msg.get("body", {}).get("content", "")
            content_type = msg.get("body", {}).get("contentType", "text")
            parsed = _parse_quote_email(subject, body_content, content_type)

            con.execute("""
            INSERT INTO lam_quotes
                (date_received, sent_to, subject, job_name, customer, location,
                 product, price, quantities, amount, est_freight, lead_time)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                date_received, sent_to, subject,
                parsed.get('job_name'), parsed.get('customer'), parsed.get('location'),
                parsed.get('product'), parsed.get('price'), parsed.get('quantities'),
                parsed.get('amount'), parsed.get('est_freight'), parsed.get('lead_time'),
            ))
            inserted += 1

    parts = []
    if inserted: parts.append(f"{inserted} new quote{'' if inserted==1 else 's'} imported")
    if updated:  parts.append(f"{updated} updated")
    msg_text = "✓ " + ", ".join(parts) if parts else "✓ Already up to date"
    return {"inserted": inserted, "skipped": skipped, "message": msg_text}

@app.get("/api/hydrotech-dashboard")
def hydrotech_dashboard():
    with get_db() as con:
        totals = con.execute("""
            SELECT
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total_amount,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won_amount,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal_amount,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost_amount
            FROM hydrotech_quotes
        """).fetchone()
        by_loc = con.execute("""
            SELECT location,
                COUNT(*) as count,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal
            FROM hydrotech_quotes WHERE location IS NOT NULL
            GROUP BY location ORDER BY total DESC
        """).fetchall()
        by_month = con.execute("""
            SELECT substr(date_received,1,7) as month,
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count
            FROM hydrotech_quotes
            WHERE date_received IS NOT NULL AND date_received != ''
              AND substr(date_received,1,7) >= substr(date('now','-11 months'),1,7)
            GROUP BY month ORDER BY month
        """).fetchall()
        by_status = con.execute("""
            SELECT status, COUNT(*) as count, COALESCE(SUM(amount),0) as amount
            FROM hydrotech_quotes WHERE status IS NOT NULL
            GROUP BY status ORDER BY amount DESC
        """).fetchall()
        locations = [r[0] for r in con.execute(
            "SELECT DISTINCT location FROM hydrotech_quotes WHERE location IS NOT NULL ORDER BY location"
        ).fetchall()]
        raw_close_ht = con.execute(
            "SELECT close_date, status, amount FROM hydrotech_quotes "
            "WHERE (deleted IS NULL OR deleted=0) AND close_date IS NOT NULL AND close_date != ''"
        ).fetchall()
        from collections import defaultdict as _dd2
        close_acc_ht = _dd2(lambda: dict(total=0.0, won=0.0, verbal=0.0, open=0.0))
        for row in raw_close_ht:
            dt = parse_date(row["close_date"])
            if not dt: continue
            ym = dt.strftime('%Y-%m')
            amt = float(row["amount"] or 0)
            st = (row["status"] or '').strip()
            if st in ('Lost', 'Duplicate'): continue
            close_acc_ht[ym]['total'] += amt
            if st == 'Won':     close_acc_ht[ym]['won']    += amt
            elif st == 'Verbal': close_acc_ht[ym]['verbal'] += amt
            else:                close_acc_ht[ym]['open']   += amt
        by_close_month_ht = [{'month': k, 'total': round(v['total'], 2), 'won': round(v['won'], 2),
                               'verbal': round(v['verbal'], 2), 'open': round(v['open'], 2)}
                              for k, v in sorted(close_acc_ht.items())]
        return {
            "totals": dict(totals),
            "by_location": [dict(r) for r in by_loc],
            "by_month": [dict(r) for r in by_month],
            "by_status": [dict(r) for r in by_status],
            "locations": locations,
            "by_close_month": by_close_month_ht,
        }

@app.get("/api/lam-dashboard")
def lam_dashboard():
    with get_db() as con:
        totals = con.execute("""
            SELECT
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total_amount,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won_amount,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal_amount,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost_amount
            FROM lam_quotes WHERE (deleted IS NULL OR deleted=0)
        """).fetchone()
        by_loc = con.execute("""
            SELECT location,
                COUNT(*) as count,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal
            FROM lam_quotes WHERE (deleted IS NULL OR deleted=0) AND location IS NOT NULL
            GROUP BY location ORDER BY total DESC
        """).fetchall()
        by_month = con.execute("""
            SELECT substr(date_received,1,7) as month,
                COUNT(*) as total_quotes,
                COALESCE(SUM(amount),0) as total,
                COALESCE(SUM(CASE WHEN status='Won'    THEN amount ELSE 0 END),0) as won,
                COALESCE(SUM(CASE WHEN status='Verbal' THEN amount ELSE 0 END),0) as verbal,
                COALESCE(SUM(CASE WHEN status='Lost'   THEN amount ELSE 0 END),0) as lost,
                COUNT(CASE WHEN status='Won'    THEN 1 END) as won_count,
                COUNT(CASE WHEN status='Verbal' THEN 1 END) as verbal_count,
                COUNT(CASE WHEN status='Lost'   THEN 1 END) as lost_count
            FROM lam_quotes
            WHERE (deleted IS NULL OR deleted=0) AND date_received IS NOT NULL AND date_received != ''
              AND substr(date_received,1,7) >= substr(date('now','-11 months'),1,7)
            GROUP BY month ORDER BY month
        """).fetchall()
        by_status = con.execute("""
            SELECT status, COUNT(*) as count, COALESCE(SUM(amount),0) as amount
            FROM lam_quotes WHERE (deleted IS NULL OR deleted=0) AND status IS NOT NULL
            GROUP BY status ORDER BY amount DESC
        """).fetchall()
        raw_close_lam = con.execute(
            "SELECT close_date, status, amount FROM lam_quotes "
            "WHERE (deleted IS NULL OR deleted=0) AND close_date IS NOT NULL AND close_date != ''"
        ).fetchall()
        from collections import defaultdict as _dd3
        close_acc_lam = _dd3(lambda: dict(total=0.0, won=0.0, verbal=0.0, open=0.0))
        for row in raw_close_lam:
            dt = parse_date(row["close_date"])
            if not dt: continue
            ym = dt.strftime('%Y-%m')
            amt = float(row["amount"] or 0)
            st = (row["status"] or '').strip()
            if st in ('Lost', 'Duplicate'): continue
            close_acc_lam[ym]['total'] += amt
            if st == 'Won':      close_acc_lam[ym]['won']    += amt
            elif st == 'Verbal': close_acc_lam[ym]['verbal'] += amt
            else:                close_acc_lam[ym]['open']   += amt
        by_close_month_lam = [{'month': k, 'total': round(v['total'], 2), 'won': round(v['won'], 2),
                               'verbal': round(v['verbal'], 2), 'open': round(v['open'], 2)}
                              for k, v in sorted(close_acc_lam.items())]
        return {
            "totals": dict(totals),
            "by_location": [dict(r) for r in by_loc],
            "by_month": [dict(r) for r in by_month],
            "by_status": [dict(r) for r in by_status],
            "by_close_month": by_close_month_lam,
        }


# ── Contacts ─────────────────────────────────────────────────────────────────

import re as _re

def _strip_html(html: str) -> str:
    html = _re.sub(r'<style[^>]*>.*?</style>', '', html, flags=_re.DOTALL|_re.IGNORECASE)
    html = _re.sub(r'<script[^>]*>.*?</script>', '', html, flags=_re.DOTALL|_re.IGNORECASE)
    html = _re.sub(r'<br\s*/?>', '\n', html, flags=_re.IGNORECASE)
    html = _re.sub(r'</?(?:p|div|tr|td|li|h\d)[^>]*>', '\n', html, flags=_re.IGNORECASE)
    html = _re.sub(r'<[^>]+>', '', html)
    return (html.replace('&nbsp;', ' ').replace('&amp;', '&')
                .replace('&lt;', '<').replace('&gt;', '>').replace('&#13;', '').replace('\r', ''))

def _extract_phone(text: str) -> str:
    m = _re.search(r'(\+?1?[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4})', text)
    return m.group(1).strip() if m else ''

def _extract_sig_text(body_text: str) -> str:
    """Return the likely email signature block (last few lines before any reply chain)."""
    sep_patterns = [
        r'^-{3,}[\s\S]*?original\s+message',
        r'^on\s+.{10,200}\s+wrote:',
        r'^from:\s*\S+@',
        r'^sent\s+from\s+my',
        r'^_{5,}',
    ]
    lines = body_text.split('\n')
    cutoff = len(lines)
    for i, line in enumerate(lines):
        for pat in sep_patterns:
            if _re.match(pat, line.strip(), _re.IGNORECASE):
                cutoff = i
                break
        if cutoff < len(lines):
            break
    non_empty = [l for l in lines[:cutoff] if l.strip()]
    if len(non_empty) <= 1:
        return ''
    return '\n'.join(non_empty[-8:])

def init_contacts_db():
    with get_db() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS contacts (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT,
            company         TEXT,
            phone           TEXT,
            email           TEXT UNIQUE,
            location        TEXT,
            product_line    TEXT,
            customer_type   TEXT,
            notes           TEXT,
            manually_edited INTEGER DEFAULT 0,
            created_at      TEXT DEFAULT (datetime('now')),
            updated_at      TEXT DEFAULT (datetime('now'))
        )""")
        # Add columns to existing tables that predate this change
        for col, defn in [
            ("manually_edited", "INTEGER DEFAULT 0"),
            ("customer_type", "TEXT"),
            ("region", "TEXT"),
        ]:
            try:
                con.execute(f"ALTER TABLE contacts ADD COLUMN {col} {defn}")
            except Exception:
                pass  # Column already exists
        con.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            id                        INTEGER PRIMARY KEY AUTOINCREMENT,
            name                      TEXT NOT NULL,
            address                   TEXT,
            phone                     TEXT,
            website                   TEXT,
            region                    TEXT,
            notes                     TEXT,
            strong_market_partner     INTEGER DEFAULT 0,
            large_account_opportunity INTEGER DEFAULT 0,
            created_at                TEXT DEFAULT (datetime('now')),
            updated_at                TEXT DEFAULT (datetime('now'))
        )""")
        # Migrations for companies columns added after initial deploy
        for col, defn in [
            ("strong_market_partner",     "INTEGER DEFAULT 0"),
            ("large_account_opportunity", "INTEGER DEFAULT 0"),
            ("region",                    "TEXT"),
            ("account_type",              "TEXT DEFAULT ''"),
        ]:
            try:
                con.execute(f"ALTER TABLE companies ADD COLUMN {col} {defn}")
            except Exception:
                pass
        con.execute("""
        CREATE TABLE IF NOT EXISTS company_contacts (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id  INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
            contact_id  INTEGER NOT NULL REFERENCES contacts(id) ON DELETE CASCADE,
            role        TEXT,
            UNIQUE(company_id, contact_id)
        )""")
        # Migration: switch from contact_id to company_id (drop+recreate if old schema)
        try:
            cols = [r[1] for r in con.execute("PRAGMA table_info(quote_contractors)").fetchall()]
            if 'contact_id' in cols and 'company_id' not in cols:
                con.execute("DROP TABLE IF EXISTS quote_contractors")
        except Exception:
            pass
        con.execute("""
        CREATE TABLE IF NOT EXISTS quote_contractors (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_id   INTEGER NOT NULL,
            quote_type TEXT NOT NULL,
            company_id INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
            UNIQUE(quote_id, quote_type, company_id)
        )""")

def init_users_db():
    """Create users table and seed initial accounts if they don't exist."""
    INITIAL_USERS = [
        ('mwalker@specformbc.com',   True),
        ('bbautista@specformbc.com', False),
        ('ftrevino@specformbc.com',  False),
        ('brentm@specformbc.com',    False),
    ]
    DEFAULT_PASSWORD = 'Specform2026!'
    with get_db() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            email      TEXT UNIQUE NOT NULL COLLATE NOCASE,
            pwd_hash   TEXT NOT NULL,
            is_admin   INTEGER DEFAULT 0,
            is_active  INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        )""")
        try:
            con.execute("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT NULL")
        except Exception:
            pass
        for email, is_admin in INITIAL_USERS:
            try:
                con.execute(
                    "INSERT INTO users (email, pwd_hash, is_admin) VALUES (?,?,?)",
                    (email, _hash_pw(DEFAULT_PASSWORD), 1 if is_admin else 0)
                )
            except Exception:
                pass  # Already exists

class ContactIn(BaseModel):
    name:          Optional[str] = None
    company:       Optional[str] = None
    phone:         Optional[str] = None
    email:         Optional[str] = None
    location:      Optional[str] = None
    product_line:  Optional[str] = None
    customer_type: Optional[str] = None
    region:        Optional[str] = None
    notes:         Optional[str] = None



@app.get("/api/quote-contractors")
def list_quote_contractors(quote_id: int = Query(...), quote_type: str = Query(...)):
    with get_db() as con:
        rows = con.execute("""
            SELECT qc.id, co.id as company_id, co.name, co.region, co.phone, co.website
            FROM quote_contractors qc
            JOIN companies co ON co.id = qc.company_id
            WHERE qc.quote_id=? AND qc.quote_type=?
            ORDER BY co.name
        """, (quote_id, quote_type)).fetchall()
        return [dict(r) for r in rows]

class QuoteContractorIn(BaseModel):
    quote_id: int
    quote_type: str
    company_id: int

@app.post("/api/quote-contractors", status_code=201)
def add_quote_contractor(data: QuoteContractorIn):
    with get_db() as con:
        try:
            cur = con.execute(
                "INSERT INTO quote_contractors (quote_id, quote_type, company_id) VALUES (?,?,?)",
                (data.quote_id, data.quote_type, data.company_id)
            )
            return {"id": cur.lastrowid}
        except Exception:
            raise HTTPException(409, "Already linked")

@app.delete("/api/quote-contractors/{entry_id}")
def remove_quote_contractor(entry_id: int):
    with get_db() as con:
        con.execute("DELETE FROM quote_contractors WHERE id=?", (entry_id,))
        return {"ok": True}

@app.get("/api/architects")
def list_architects():
    with get_db() as con:
        rows = con.execute(
            "SELECT id, name, company, phone, email, location, region FROM contacts "
            "WHERE customer_type='Architect' ORDER BY name"
        ).fetchall()
        return [dict(r) for r in rows]

@app.get("/api/contacts")
def list_contacts(
    product_line: Optional[str] = Query(None),
    location:     Optional[str] = Query(None),
    search:       Optional[str] = Query(None),
):
    with get_db() as con:
        sql = """
            SELECT ct.*,
                   co.id   AS linked_company_id,
                   co.name AS linked_company_name
            FROM contacts ct
            LEFT JOIN (
                SELECT contact_id, MIN(company_id) AS company_id
                FROM company_contacts GROUP BY contact_id
            ) cc ON cc.contact_id = ct.id
            LEFT JOIN companies co ON co.id = cc.company_id
            WHERE 1=1
        """
        params = []
        if product_line and product_line != 'All':
            sql += " AND (ct.product_line = ? OR ct.product_line = 'Both')"
            params.append(product_line)
        if location and location != 'All':
            sql += " AND ct.location = ?"
            params.append(location)
        if search:
            sql += " AND (ct.name LIKE ? OR ct.company LIKE ? OR ct.email LIKE ? OR ct.location LIKE ? OR ct.phone LIKE ?)"
            s = f"%{search}%"
            params += [s, s, s, s, s]
        sql += " ORDER BY COALESCE(NULLIF(ct.location,''),'zzz'), COALESCE(NULLIF(ct.name,''),'zzz'), ct.email"
        return [dict(r) for r in con.execute(sql, params).fetchall()]

@app.post("/api/contacts", status_code=201)
def create_contact(c: ContactIn):
    with get_db() as con:
        try:
            cur = con.execute("""
            INSERT INTO contacts (name,company,phone,email,location,product_line,customer_type,region,notes)
            VALUES (?,?,?,?,?,?,?,?,?)
            """, (c.name,c.company,c.phone,c.email,c.location,c.product_line,c.customer_type,c.region,c.notes))
            return {"id": cur.lastrowid}
        except Exception:
            raise HTTPException(409, "A contact with this email already exists")

@app.put("/api/contacts/{contact_id}")
def update_contact(contact_id: int, c: ContactIn):
    with get_db() as con:
        existing = con.execute("SELECT id FROM contacts WHERE id=?", (contact_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Contact not found")
        con.execute("""
        UPDATE contacts SET name=?,company=?,phone=?,email=?,location=?,product_line=?,
            customer_type=?,region=?,notes=?,manually_edited=1,updated_at=datetime('now')
        WHERE id=?
        """, (c.name,c.company,c.phone,c.email,c.location,c.product_line,c.customer_type,c.region,c.notes,contact_id))
        return {"ok": True}

@app.delete("/api/contacts/{contact_id}")
def delete_contact(contact_id: int):
    with get_db() as con:
        con.execute("DELETE FROM contacts WHERE id=?", (contact_id,))
        return {"ok": True}

_EMAIL_RE = _re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

@app.post("/api/contacts/import-from-quotes")
def import_contacts_from_quotes():
    """Scan both quote tables and create contact entries for unique customers / sent_to emails."""
    from collections import defaultdict
    with get_db() as con:
        rmax_rows  = con.execute("SELECT DISTINCT sent_to, customer, location FROM quotes WHERE (deleted IS NULL OR deleted=0) AND sent_to IS NOT NULL AND sent_to != ''").fetchall()
        hydro_rows = con.execute("SELECT DISTINCT sent_to, customer, location FROM hydrotech_quotes WHERE sent_to IS NOT NULL AND sent_to != ''").fetchall()

        # Build a map keyed by email (validated) or company name (as fallback)
        contact_map = {}  # key -> {"email": str|None, "company": str, "location": str, "sources": set}

        def _add_row(row, source):
            raw    = (row["sent_to"] or "").strip()
            email  = raw.lower() if _EMAIL_RE.match(raw) else None
            if not email:
                return  # skip entries with no valid email address
            company = (row["customer"] or "").strip()
            location = (row["location"] or "").strip()
            key = email
            if key not in contact_map:
                contact_map[key] = {"email": email, "company": company, "location": location, "sources": set()}
            else:
                if email and not contact_map[key]["email"]:
                    contact_map[key]["email"] = email
                if company and not contact_map[key]["company"]:
                    contact_map[key]["company"] = company
                if location and not contact_map[key]["location"]:
                    contact_map[key]["location"] = location
            contact_map[key]["sources"].add(source)

        for row in rmax_rows:  _add_row(row, "RMAX")
        for row in hydro_rows: _add_row(row, "Hydrotech")

        inserted = updated = skipped = 0
        for key, info in contact_map.items():
            email        = info["email"]
            company      = info["company"]
            location     = info["location"]
            product_line = "Both" if len(info["sources"]) > 1 else list(info["sources"])[0]

            # Dedup by email only — same email = same person; different email = new contact
            existing = con.execute(
                "SELECT id, email, company, location, product_line, manually_edited FROM contacts WHERE email=?",
                (email,)
            ).fetchone() if email else None
            if existing:
                if existing["manually_edited"]:
                    skipped += 1
                    continue
                updates = {}
                if email    and not existing["email"]:    updates["email"]    = email
                if company  and not existing["company"]:  updates["company"]  = company
                if location and not existing["location"]: updates["location"] = location
                if product_line == "Both" and existing["product_line"] != "Both":
                    updates["product_line"] = "Both"
                if updates:
                    set_clause = ", ".join(f"{k}=?" for k in updates)
                    con.execute(f"UPDATE contacts SET {set_clause}, updated_at=datetime('now') WHERE id=?",
                                list(updates.values()) + [existing["id"]])
                    updated += 1
                else:
                    skipped += 1
            else:
                con.execute(
                    "INSERT INTO contacts (email, company, location, product_line) VALUES (?,?,?,?)",
                    (email, company, location, product_line)
                )
                inserted += 1

        return {"inserted": inserted, "updated": updated, "skipped": skipped, "total": inserted + updated + skipped}

@app.get("/api/contacts/fetch-signature/{email:path}")
def fetch_contact_signature(email: str):
    """Search inbox for an email FROM this address and parse their signature for contact info."""
    if not GRAPH_CLIENT_SECRET:
        return {"found": False, "reason": "Graph API not configured"}
    try:
        import requests as _req
    except ImportError:
        return {"found": False, "reason": "requests package not installed"}
    try:
        token = _get_graph_token()
    except Exception as e:
        return {"found": False, "reason": str(e)}

    url = (f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/messages"
           f"?$filter=from/emailAddress/address eq '{email}'"
           f"&$select=body,from,subject&$top=1"
           f"&$orderby=receivedDateTime desc")
    try:
        resp = _req.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=20)
    except Exception as e:
        return {"found": False, "reason": f"Network error: {e}"}
    if resp.status_code == 401:
        return {"found": False, "reason": "Unauthorized — verify Graph API Mail.Read permission"}
    if resp.status_code != 200:
        return {"found": False, "reason": f"Graph API returned {resp.status_code}"}

    msgs = resp.json().get("value", [])
    if not msgs:
        return {"found": False, "reason": "No emails found from this address in your inbox"}

    msg = msgs[0]
    body_html = msg.get("body", {}).get("content", "")
    content_type = msg.get("body", {}).get("contentType", "text")
    sender_name  = msg.get("from", {}).get("emailAddress", {}).get("name", "")

    body_text = _strip_html(body_html) if content_type == "html" else body_html
    body_text = _re.sub(r'\n{3,}', '\n\n', body_text).strip()

    sig_text = _extract_sig_text(body_text)
    phone    = _extract_phone(sig_text or body_text)

    return {
        "found": True,
        "name":           sender_name,
        "phone":          phone,
        "signature_text": sig_text,
    }

@app.post("/api/contacts/scan-outlook-folders")
def scan_outlook_folders(folders: Optional[str] = Query(None)):
    """
    Scan one or more Outlook mail folders for email signatures and bulk-populate contacts.
    folders = comma-separated folder names, defaults to 'RMAX Quotes,Hydrotech Quotes'
    """
    if not GRAPH_CLIENT_SECRET:
        return {"error": "Graph API not configured — add GRAPH_CLIENT_SECRET to Railway env vars"}
    try:
        import requests as _req
    except ImportError:
        return {"error": "requests package not installed"}
    try:
        token = _get_graph_token()
    except Exception as e:
        return {"error": str(e)}

    folder_names = [f.strip() for f in (folders or "RMAX Quotes,Hydrotech Quotes").split(",") if f.strip()]
    headers = {"Authorization": f"Bearer {token}"}

    def _find_folder_id(name: str) -> str | None:
        """Look for a mail folder by display name — top-level and inbox children."""
        for url in [
            f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders?$filter=displayName eq '{name}'&$select=id",
            f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}/mailFolders/Inbox/childFolders?$filter=displayName eq '{name}'&$select=id",
        ]:
            try:
                r = _req.get(url, headers=headers, timeout=20)
                if r.status_code == 200:
                    vals = r.json().get("value", [])
                    if vals:
                        return vals[0]["id"]
            except Exception:
                pass
        return None

    def _paginate_messages(url_start):
        """Fetch up to 1000 messages from a paginated Graph URL."""
        msgs = []
        next_url = url_start
        while next_url and len(msgs) < 1000:
            try:
                r = _req.get(next_url, headers=headers, timeout=30)
                if r.status_code != 200:
                    break
                data = r.json()
                msgs.extend(data.get("value", []))
                next_url = data.get("@odata.nextLink")
            except Exception:
                break
        return msgs

    results = []
    total_created = total_updated = total_skipped = 0

    with get_db() as con:
        for folder_name in folder_names:
            pl = "Hydrotech" if "hydrotech" in folder_name.lower() else "RMAX"

            # Collect messages from named folder (if it exists) + Sent Items filtered by subject
            msgs = []
            folder_id = _find_folder_id(folder_name)
            if folder_id:
                msgs = _paginate_messages(
                    f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}"
                    f"/mailFolders/{folder_id}/messages"
                    f"?$select=from,body,subject,toRecipients&$top=100"
                )

            # Also scan Sent Items for emails with this folder name in the subject
            # (catches quotes that never got moved to the folder)
            subject_kw = folder_name.replace("'", "''")  # escape single quotes for OData
            sent_msgs = _paginate_messages(
                f"https://graph.microsoft.com/v1.0/users/{GRAPH_USER}"
                f"/mailFolders/SentItems/messages"
                f"?$filter=contains(subject,'{subject_kw}')"
                f"&$select=from,body,subject,toRecipients&$top=100"
            )
            # Merge, deduplicating by message id
            seen_ids = {m.get("id") for m in msgs}
            for m in sent_msgs:
                if m.get("id") not in seen_ids:
                    msgs.append(m)
                    seen_ids.add(m.get("id"))

            if not msgs:
                results.append({"folder": folder_name, "error": "No messages found in folder or Sent Items"})
                continue

            created = updated = skipped = 0
            seen = set()

            for msg in msgs:
                # Quote was sent TO the customer — read toRecipients for their address
                recipients = msg.get("toRecipients", [])
                if not recipients:
                    continue
                contact_addr = recipients[0].get("emailAddress", {})
                email = (contact_addr.get("address") or "").strip().lower()
                name  = (contact_addr.get("name")    or "").strip()
                if not email or email in seen:
                    continue
                seen.add(email)

                body_html    = msg.get("body", {}).get("content", "")
                content_type = msg.get("body", {}).get("contentType", "text")
                body_text    = _strip_html(body_html) if content_type == "html" else body_html
                body_text    = _re.sub(r'\n{3,}', '\n\n', body_text).strip()
                sig_text     = _extract_sig_text(body_text)
                phone        = _extract_phone(sig_text or body_text)

                # Pull company/location from quotes tables
                qrow = (con.execute("SELECT customer, location FROM quotes WHERE LOWER(sent_to)=? AND customer!='' LIMIT 1", (email,)).fetchone()
                        or con.execute("SELECT customer, location FROM hydrotech_quotes WHERE LOWER(sent_to)=? AND customer!='' LIMIT 1", (email,)).fetchone())
                company  = (qrow["customer"] if qrow else "") or ""
                location = (qrow["location"] if qrow else "") or ""

                # Dedup by email only — same email = same person; different email = new contact
                existing = con.execute("SELECT * FROM contacts WHERE email=?", (email,)).fetchone()
                if existing:
                    # Never touch a contact the user has manually edited
                    if existing["manually_edited"]:
                        skipped += 1
                        continue
                    updates = {}
                    if name     and not existing["name"]:     updates["name"]     = name
                    if phone    and not existing["phone"]:    updates["phone"]    = phone
                    if company  and not existing["company"]:  updates["company"]  = company
                    if location and not existing["location"]: updates["location"] = location
                    # Fill in email if matched by company name and email was blank
                    if email and not existing["email"]:       updates["email"]    = email
                    # Upgrade product_line to Both if contact now appears in a second line
                    if existing["product_line"] and existing["product_line"] != pl and existing["product_line"] != "Both":
                        updates["product_line"] = "Both"
                    if updates:
                        set_clause = ", ".join(f"{k}=?" for k in updates)
                        con.execute(
                            f"UPDATE contacts SET {set_clause}, updated_at=datetime('now') WHERE id=?",
                            list(updates.values()) + [existing["id"]]
                        )
                        updated += 1
                    else:
                        skipped += 1
                else:
                    con.execute(
                        "INSERT INTO contacts (name,email,phone,company,location,product_line) VALUES (?,?,?,?,?,?)",
                        (name, email, phone, company, location, pl)
                    )
                    created += 1

            total_created += created
            total_updated += updated
            total_skipped += skipped
            results.append({
                "folder":           folder_name,
                "messages_scanned": len(msgs),
                "unique_senders":   len(seen),
                "created":          created,
                "updated":          updated,
                "skipped":          skipped,
            })

    return {
        "folders":       results,
        "total_created": total_created,
        "total_updated": total_updated,
        "total_skipped": total_skipped,
    }

# ── Company Accounts ──────────────────────────────────────────────────────────
class CompanyIn(BaseModel):
    name:                      str
    address:                   Optional[str] = None
    phone:                     Optional[str] = None
    website:                   Optional[str] = None
    region:                    Optional[str] = None
    notes:                     Optional[str] = None
    strong_market_partner:     Optional[int] = 0
    large_account_opportunity: Optional[int] = 0
    account_type:              Optional[str] = None

@app.get("/api/companies")
def list_companies(search: Optional[str] = Query(None), region: Optional[str] = Query(None), account_type: Optional[str] = Query(None)):
    with get_db() as con:
        sql = """
            SELECT c.*,
                   COUNT(DISTINCT cc.contact_id) as contact_count,
                   COALESCE((SELECT SUM(amount) FROM quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0) AS rmax_quoted,
                   COALESCE((SELECT SUM(amount) FROM hydrotech_quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0) AS hydrotech_quoted,
                   COALESCE((SELECT SUM(amount) FROM glassworks_quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0) AS glassworks_quoted,
                   COALESCE((SELECT SUM(amount) FROM quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0)
                 + COALESCE((SELECT SUM(amount) FROM hydrotech_quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0)
                 + COALESCE((SELECT SUM(amount) FROM glassworks_quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0)
                   AS total_quoted
            FROM companies c
            LEFT JOIN company_contacts cc ON cc.company_id = c.id
            WHERE 1=1
        """
        params = []
        if search:
            sql += " AND (c.name LIKE ? OR c.phone LIKE ? OR c.address LIKE ?)"
            s = f"%{search}%"
            params += [s, s, s]
        if region and region != 'All':
            sql += " AND c.region = ?"
            params.append(region)
        if account_type:
            sql += " AND c.account_type = ?"
            params.append(account_type)
        sql += " GROUP BY c.id ORDER BY c.name COLLATE NOCASE"
        return [dict(r) for r in con.execute(sql, params).fetchall()]

@app.post("/api/companies", status_code=201)
def create_company(c: CompanyIn):
    with get_db() as con:
        cur = con.execute(
            "INSERT INTO companies (name,address,phone,website,region,notes,strong_market_partner,large_account_opportunity,account_type) VALUES (?,?,?,?,?,?,?,?,?)",
            (c.name, c.address, c.phone, c.website, c.region, c.notes, c.strong_market_partner or 0, c.large_account_opportunity or 0, c.account_type or '')
        )
        return {"id": cur.lastrowid}

@app.put("/api/companies/{company_id}")
def update_company(company_id: int, c: CompanyIn):
    with get_db() as con:
        existing = con.execute("SELECT id FROM companies WHERE id=?", (company_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Company not found")
        con.execute("""
            UPDATE companies SET name=?,address=?,phone=?,website=?,region=?,notes=?,
                strong_market_partner=?,large_account_opportunity=?,account_type=?,updated_at=datetime('now')
            WHERE id=?
        """, (c.name, c.address, c.phone, c.website, c.region, c.notes,
              c.strong_market_partner or 0, c.large_account_opportunity or 0, c.account_type or '', company_id))
        return {"ok": True}

@app.delete("/api/companies/{company_id}")
def delete_company(company_id: int):
    with get_db() as con:
        con.execute("DELETE FROM company_contacts WHERE company_id=?", (company_id,))
        con.execute("DELETE FROM companies WHERE id=?", (company_id,))
        return {"ok": True}

@app.post("/api/companies/import-from-quotes")
def import_companies_from_quotes():
    """Scan all quote tables and create Company Account entries from unique customer names,
    carrying over the most-used region from that customer's quotes."""
    from collections import defaultdict, Counter
    with get_db() as con:
        rows = []
        for table in ("quotes", "hydrotech_quotes", "glassworks_quotes"):
            try:
                rs = con.execute(
                    f"SELECT customer, region FROM {table} "
                    f"WHERE (deleted IS NULL OR deleted=0) "
                    f"AND customer IS NOT NULL AND TRIM(customer) != ''"
                ).fetchall()
                rows.extend(rs)
            except Exception:
                pass

        # Group all regions seen per customer name (case-insensitive key)
        customer_map = defaultdict(lambda: {"canonical": "", "regions": []})
        for row in rows:
            name = (row["customer"] or "").strip()
            region = (row["region"] or "").strip()
            if not name:
                continue
            key = name.lower()
            if not customer_map[key]["canonical"]:
                customer_map[key]["canonical"] = name
            if region:
                customer_map[key]["regions"].append(region)

        inserted = updated = skipped = 0
        for key, info in customer_map.items():
            name = info["canonical"]
            non_empty = info["regions"]
            best_region = Counter(non_empty).most_common(1)[0][0] if non_empty else None

            existing = con.execute(
                "SELECT id, region FROM companies WHERE LOWER(name) = ?", (key,)
            ).fetchone()

            if existing:
                if best_region and not existing["region"]:
                    con.execute(
                        "UPDATE companies SET region=?, updated_at=datetime('now') WHERE id=?",
                        (best_region, existing["id"])
                    )
                    updated += 1
                else:
                    skipped += 1
            else:
                con.execute(
                    "INSERT INTO companies (name, region) VALUES (?, ?)",
                    (name, best_region)
                )
                inserted += 1

        return {"inserted": inserted, "updated": updated, "skipped": skipped,
                "total": inserted + updated + skipped}

@app.get("/api/companies/{company_id}/contacts")
def list_company_contacts(company_id: int):
    with get_db() as con:
        rows = con.execute("""
            SELECT ct.*, cc.role, cc.id as link_id
            FROM contacts ct
            JOIN company_contacts cc ON cc.contact_id = ct.id
            WHERE cc.company_id = ?
            ORDER BY COALESCE(NULLIF(ct.name,''),'zzz')
        """, (company_id,)).fetchall()
        return [dict(r) for r in rows]

class CompanyContactIn(BaseModel):
    contact_id: int
    role:       Optional[str] = None

@app.post("/api/companies/{company_id}/contacts", status_code=201)
def add_company_contact(company_id: int, body: CompanyContactIn):
    with get_db() as con:
        existing = con.execute("SELECT id FROM companies WHERE id=?", (company_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Company not found")
        try:
            con.execute(
                "INSERT INTO company_contacts (company_id,contact_id,role) VALUES (?,?,?)",
                (company_id, body.contact_id, body.role)
            )
        except Exception:
            raise HTTPException(409, "Contact already linked to this company")
        return {"ok": True}

@app.delete("/api/companies/{company_id}/contacts/{contact_id}")
def remove_company_contact(company_id: int, contact_id: int):
    with get_db() as con:
        con.execute(
            "DELETE FROM company_contacts WHERE company_id=? AND contact_id=?",
            (company_id, contact_id)
        )
        return {"ok": True}

@app.put("/api/companies/{company_id}/contacts/{contact_id}")
def update_company_contact_role(company_id: int, contact_id: int, body: CompanyContactIn):
    with get_db() as con:
        con.execute(
            "UPDATE company_contacts SET role=? WHERE company_id=? AND contact_id=?",
            (body.role, company_id, contact_id)
        )
        return {"ok": True}

@app.get("/api/companies/{company_id}/quote-stats")
def company_quote_stats(company_id: int, period: str = Query("all")):
    """Sum amount quoted and amount won across all 3 quote tables for a company.
    period: 'all' | 'ytd' | 'YYYY-MM'
    """
    import datetime
    tables = [
        ("quotes",           "deleted"),
        ("hydrotech_quotes", "deleted"),
        ("glassworks_quotes","deleted"),
    ]
    # Build date clause
    now = datetime.date.today()
    date_clause = ""
    date_params: list = []
    if period == "ytd":
        date_clause = "AND date_received >= ?"
        date_params = [f"{now.year}-01-01"]
    elif len(period) == 7 and period[4] == "-":   # YYYY-MM
        date_clause = "AND date_received LIKE ?"
        date_params = [f"{period}%"]

    total_quoted = 0.0
    total_won    = 0.0
    quote_count  = 0
    won_count    = 0
    with get_db() as con:
        for table, del_col in tables:
            cols = {r[1] for r in con.execute(f"PRAGMA table_info({table})")}
            if "company_id" not in cols or "amount" not in cols:
                continue
            where_del = f"AND ({del_col} IS NULL OR {del_col}=0)" if del_col in cols else ""
            has_date  = "date_received" in cols
            d_clause  = date_clause if has_date else ""
            params    = [company_id] + (date_params if has_date else [])
            rows = con.execute(
                f"SELECT amount, status FROM {table} WHERE company_id=? {where_del} {d_clause}",
                params
            ).fetchall()
            for r in rows:
                amt = r["amount"] or 0.0
                total_quoted += amt
                quote_count  += 1
                if r["status"] == "Won":
                    total_won += amt
                    won_count += 1
    return {
        "total_quoted": total_quoted,
        "total_won":    total_won,
        "quote_count":  quote_count,
        "won_count":    won_count,
        "period":       period,
    }

@app.get("/api/companies/export")
def export_companies():
    """Export company accounts + contacts: product line → region → company (bold) → contacts (sub-rows)."""
    import io
    import openpyxl
    from collections import defaultdict
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    thin      = Side(style="thin", color="D1D5DB")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
    WRAP      = Alignment(wrap_text=True, vertical="top")
    LEFT      = Alignment(horizontal="left", vertical="center")
    CTR       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    INDENTED  = Alignment(horizontal="left", vertical="center", indent=1)

    PL_FILLS = {
        "RMAX":       PatternFill("solid", fgColor="1F4E79"),
        "Hydrotech":  PatternFill("solid", fgColor="064E3B"),
        "Glassworks": PatternFill("solid", fgColor="7C2D12"),
    }
    PL_FONT     = Font(bold=True, color="FFFFFF", size=12)
    REGION_FILL = PatternFill("solid", fgColor="DBEAFE")
    REGION_FONT = Font(bold=True, color="1E3A5F", size=10)
    CO_FONT     = Font(bold=True, size=10)
    CO_ALT_FILL = PatternFill("solid", fgColor="F4F6F8")
    CT_FONT     = Font(italic=True, size=9, color="475569")
    CT_FILL     = PatternFill("solid", fgColor="F8FAFC")
    HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT    = Font(bold=True, color="FFFFFF", size=11)
    MONEY_FMT   = '#,##0.00'
    NCOLS       = 10

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Accounts"

    # Column header row
    # Col 5 doubles as "SMP" for company rows and "Role" for contact sub-rows
    HEADERS = [
        "Name", "Address / Email", "Phone", "Region",
        "SMP / Role", "Large Acct Opp",
        "RMAX Quoted ($)", "Hydrotech Quoted ($)", "Glassworks Quoted ($)", "Total Quoted ($)"
    ]
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CTR, BORDER
    ws.row_dimensions[1].height = 22

    with get_db() as con:
        companies = con.execute("""
            SELECT c.*,
                   COALESCE((SELECT SUM(amount) FROM quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0) AS rmax_quoted,
                   COALESCE((SELECT SUM(amount) FROM hydrotech_quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0) AS hydrotech_quoted,
                   COALESCE((SELECT SUM(amount) FROM glassworks_quotes
                             WHERE company_id=c.id AND (deleted IS NULL OR deleted=0)),0) AS glassworks_quoted
            FROM companies c
            GROUP BY c.id
        """).fetchall()

        contacts_rows = con.execute("""
            SELECT cc.company_id, ct.name, ct.email, ct.phone, cc.role
            FROM company_contacts cc
            JOIN contacts ct ON ct.id = cc.contact_id
            ORDER BY ct.name COLLATE NOCASE
        """).fetchall()

    contacts_by_co = defaultdict(list)
    for ct in contacts_rows:
        contacts_by_co[ct["company_id"]].append(ct)

    PRODUCT_LINES = [
        ("RMAX",       "rmax_quoted"),
        ("Hydrotech",  "hydrotech_quoted"),
        ("Glassworks", "glassworks_quoted"),
    ]

    cur_row = 2

    for pl_name, pl_field in PRODUCT_LINES:
        pl_companies = [co for co in companies if (co[pl_field] or 0) > 0]
        if not pl_companies:
            continue

        region_map = defaultdict(list)
        for co in pl_companies:
            region_map[co["region"] or ""].append(co)
        regions = sorted(region_map.keys(), key=lambda r: r.lower() if r else "zzz")

        # ── Product line header row ────────────────────────────────────────────
        pl_fill = PL_FILLS[pl_name]
        for col in range(1, NCOLS + 1):
            cell = ws.cell(row=cur_row, column=col, value=pl_name if col == 1 else None)
            cell.fill, cell.border = pl_fill, BORDER
            if col == 1:
                cell.font, cell.alignment = PL_FONT, LEFT
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        for region in regions:
            region_cos = sorted(region_map[region], key=lambda c: -(c[pl_field] or 0))
            region_label = region if region else "(No Region)"

            # ── Region sub-header row ──────────────────────────────────────────
            for col in range(1, NCOLS + 1):
                cell = ws.cell(row=cur_row, column=col, value=region_label if col == 1 else None)
                cell.fill, cell.border = REGION_FILL, BORDER
                if col == 1:
                    cell.font, cell.alignment = REGION_FONT, INDENTED
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

            for idx, co in enumerate(region_cos):
                total = (co["rmax_quoted"] or 0) + (co["hydrotech_quoted"] or 0) + (co["glassworks_quoted"] or 0)
                co_fill = CO_ALT_FILL if idx % 2 == 1 else PatternFill()
                co_vals = [
                    co["name"] or "",
                    co["address"] or "",
                    co["phone"] or "",
                    co["region"] or "",
                    "Yes" if co["strong_market_partner"] else "No",
                    "Yes" if co["large_account_opportunity"] else "No",
                    round(co["rmax_quoted"] or 0, 2),
                    round(co["hydrotech_quoted"] or 0, 2),
                    round(co["glassworks_quoted"] or 0, 2),
                    round(total, 2),
                ]
                for col, v in enumerate(co_vals, 1):
                    cell = ws.cell(row=cur_row, column=col, value=v)
                    cell.fill, cell.font, cell.border, cell.alignment = co_fill, CO_FONT, BORDER, WRAP
                    if col in (7, 8, 9, 10):
                        cell.number_format = MONEY_FMT
                ws.row_dimensions[cur_row].height = 16
                cur_row += 1

                # ── Contact sub-rows (immediately below company) ───────────────
                for ct in contacts_by_co.get(co["id"], []):
                    ct_vals = [
                        f"  → {ct['name'] or ''}",  # → Contact Name
                        ct["email"] or "",
                        ct["phone"] or "",
                        "",
                        ct["role"] or "",
                        "", "", "", "", "",
                    ]
                    for col, v in enumerate(ct_vals, 1):
                        cell = ws.cell(row=cur_row, column=col, value=v)
                        cell.fill, cell.font, cell.border, cell.alignment = CT_FILL, CT_FONT, BORDER, WRAP
                    ws.row_dimensions[cur_row].height = 14
                    cur_row += 1

    ws.freeze_panes = "A2"
    col_widths = [30, 35, 15, 15, 18, 14, 16, 18, 18, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="company_accounts.xlsx"'}
    )

@app.get("/api/companies/{company_id}/quotes")
def company_quotes_list(company_id: int):
    """All quotes linked to this company across all 3 product lines."""
    tables = [
        ("quotes",            "RMAX",       "deleted"),
        ("hydrotech_quotes",  "Hydrotech",  "deleted"),
        ("glassworks_quotes", "Glassworks", "deleted"),
    ]
    results = []
    with get_db() as con:
        for table, source, del_col in tables:
            cols = {r[1] for r in con.execute(f"PRAGMA table_info({table})")}
            if "company_id" not in cols:
                continue
            where_del = f"AND ({del_col} IS NULL OR {del_col}=0)" if del_col in cols else ""
            rows = con.execute(
                f"SELECT id, date_received, job_name, subject, customer, status, amount, region "
                f"FROM {table} WHERE company_id=? {where_del} ORDER BY date_received DESC",
                (company_id,)
            ).fetchall()
            for r in rows:
                d = dict(r)
                d["source"] = source
                results.append(d)
    results.sort(key=lambda r: r.get("date_received") or "", reverse=True)
    return results

# ── Auth endpoints ────────────────────────────────────────────────────────────


class LoginIn(BaseModel):
    email:    str
    password: str

class UserIn(BaseModel):
    email:    str
    password: str
    is_admin: Optional[int] = 0

class UserUpdateIn(BaseModel):
    is_active:   Optional[int] = None
    is_admin:    Optional[int] = None
    password:    Optional[str] = None
    permissions: Optional[str] = None

def _require_admin(request: Request):
    if not request.state.user.get('is_admin'):
        raise HTTPException(403, 'Admin access required')

@app.post('/api/auth/login')
def login(body: LoginIn):
    email = body.email.strip().lower()
    with get_db() as con:
        user = con.execute(
            "SELECT * FROM users WHERE LOWER(email)=? AND is_active=1", (email,)
        ).fetchone()
    if not user or not _verify_pw(body.password, user['pwd_hash']):
        raise HTTPException(401, 'Invalid email or password')
    token = jwt.encode(
        {
            'sub':      email,
            'is_admin': bool(user['is_admin']),
            'permissions': user['permissions'],
            'exp':      datetime.utcnow() + timedelta(days=JWT_EXPIRE_DAYS),
        },
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )
    return {'token': token, 'email': email, 'is_admin': bool(user['is_admin']), 'permissions': user['permissions']}

@app.get('/api/auth/me')
def get_me(request: Request):
    return {
        'email':       request.state.user['sub'],
        'is_admin':    request.state.user.get('is_admin', False),
        'permissions': request.state.user.get('permissions'),
    }

class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str

@app.post('/api/auth/change-password')
def change_password(body: ChangePasswordIn, request: Request):
    email = request.state.user['sub']
    if len(body.new_password) < 8:
        raise HTTPException(400, 'New password must be at least 8 characters')
    with get_db() as con:
        user = con.execute("SELECT * FROM users WHERE LOWER(email)=?", (email,)).fetchone()
        if not user or not _verify_pw(body.current_password, user['pwd_hash']):
            raise HTTPException(401, 'Current password is incorrect')
        con.execute("UPDATE users SET pwd_hash=? WHERE LOWER(email)=?",
                    (_hash_pw(body.new_password), email))
    return {'ok': True}

@app.get('/api/admin/users')
def admin_list_users(request: Request):
    _require_admin(request)
    with get_db() as con:
        rows = con.execute(
            "SELECT id, email, is_admin, is_active, permissions, created_at FROM users ORDER BY email"
        ).fetchall()
    return [dict(r) for r in rows]

@app.post('/api/admin/users', status_code=201)
def admin_add_user(request: Request, body: UserIn):
    _require_admin(request)
    with get_db() as con:
        try:
            con.execute(
                "INSERT INTO users (email, pwd_hash, is_admin) VALUES (?,?,?)",
                (body.email.strip().lower(), _hash_pw(body.password), body.is_admin or 0)
            )
        except Exception:
            raise HTTPException(409, 'An account with that email already exists')
    return {'ok': True}

@app.put('/api/admin/users/{user_id}')
def admin_update_user(user_id: int, request: Request, body: UserUpdateIn):
    _require_admin(request)
    with get_db() as con:
        if not con.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone():
            raise HTTPException(404, 'User not found')
        if body.is_active is not None:
            con.execute("UPDATE users SET is_active=? WHERE id=?", (body.is_active, user_id))
        if body.is_admin is not None:
            con.execute("UPDATE users SET is_admin=? WHERE id=?", (body.is_admin, user_id))
        if body.password:
            con.execute("UPDATE users SET pwd_hash=? WHERE id=?",
                        (_hash_pw(body.password), user_id))
        if body.permissions is not None:
            con.execute("UPDATE users SET permissions=? WHERE id=?", (body.permissions, user_id))
    return {'ok': True}

@app.delete('/api/admin/users/{user_id}')
def admin_delete_user(user_id: int, request: Request):
    _require_admin(request)
    with get_db() as con:
        con.execute("DELETE FROM users WHERE id=?", (user_id,))
    return {'ok': True}

@app.post('/api/admin/users/{user_id}/send-welcome')
def admin_send_welcome(user_id: int, request: Request):
    import msal, requests as _req

    _require_admin(request)

    client_id     = os.environ.get('AZURE_CLIENT_ID', '')
    client_secret = os.environ.get('AZURE_CLIENT_SECRET', '')
    tenant_id     = os.environ.get('AZURE_TENANT_ID', GRAPH_TENANT_ID)
    from_email    = os.environ.get('FROM_EMAIL', 'mwalker@specformbc.com')
    app_url       = os.environ.get('APP_URL', 'https://specform-sales-tracker-production.up.railway.app')

    if not client_id or not client_secret:
        raise HTTPException(500, 'Email not configured — set AZURE_CLIENT_ID and AZURE_CLIENT_SECRET in Railway environment variables')

    with get_db() as con:
        user = con.execute("SELECT email FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        raise HTTPException(404, 'User not found')

    to_email = user['email']

    # Get Microsoft Graph token
    _msal_app = msal.ConfidentialClientApplication(
        client_id,
        authority=f'https://login.microsoftonline.com/{tenant_id}',
        client_credential=client_secret,
    )
    token_result = _msal_app.acquire_token_for_client(scopes=['https://graph.microsoft.com/.default'])
    if 'access_token' not in token_result:
        raise HTTPException(500, f'Email auth failed: {token_result.get("error_description", "unknown error")}')

    html = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f3f4f6;padding:40px 0;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,0.08);">
        <tr>
          <td style="background:#1F4E79;padding:32px 40px;text-align:center;">
            <h1 style="margin:0;color:#ffffff;font-size:26px;font-weight:700;letter-spacing:-0.5px;">SPECFORM SalesPartner</h1>
            <p style="margin:8px 0 0;color:rgba(255,255,255,0.75);font-size:14px;">Your sales intelligence platform</p>
          </td>
        </tr>
        <tr>
          <td style="padding:40px 40px 32px;">
            <p style="margin:0 0 16px;font-size:16px;color:#374151;line-height:1.6;">Hi there,</p>
            <p style="margin:0 0 16px;font-size:16px;color:#374151;line-height:1.6;">
              You've been added to <strong>SPECFORM SalesPartner</strong> — our internal sales tracking platform for RMAX, Hydrotech, and Glassworks quotes.
            </p>
            <p style="margin:0 0 16px;font-size:16px;color:#374151;line-height:1.6;">With SalesPartner you can:</p>
            <ul style="margin:0 0 24px 24px;padding:0;color:#374151;font-size:15px;line-height:1.8;">
              <li>Track quotes across all product lines in real time</li>
              <li>View and manage the customer directory</li>
              <li>Monitor company accounts and key opportunities</li>
              <li>See overall SPECFORM sales performance</li>
            </ul>
            <p style="margin:0 0 28px;font-size:16px;color:#374151;line-height:1.6;">
              Click below to log in with your email address (<strong>{to_email}</strong>). Your temporary password was provided separately — you can change it anytime from the app.
            </p>
            <table cellpadding="0" cellspacing="0" style="margin:0 0 32px;">
              <tr>
                <td style="background:#1F4E79;border-radius:8px;padding:14px 32px;text-align:center;">
                  <a href="{app_url}" style="color:#ffffff;font-size:16px;font-weight:700;text-decoration:none;display:inline-block;">Open SPECFORM SalesPartner →</a>
                </td>
              </tr>
            </table>
            <p style="margin:0;font-size:13px;color:#9ca3af;">
              If the button doesn't work, copy this link:<br>
              <a href="{app_url}" style="color:#1F4E79;">{app_url}</a>
            </p>
          </td>
        </tr>
        <tr>
          <td style="background:#f9fafb;border-top:1px solid #e5e7eb;padding:20px 40px;text-align:center;">
            <p style="margin:0;font-size:12px;color:#9ca3af;">Specform Building Components &nbsp;|&nbsp; Internal use only</p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body>
</html>
"""

    resp = _req.post(
        f'https://graph.microsoft.com/v1.0/users/{from_email}/sendMail',
        headers={
            'Authorization': f'Bearer {token_result["access_token"]}',
            'Content-Type': 'application/json',
        },
        json={
            'message': {
                'subject': 'Welcome to SPECFORM SalesPartner',
                'body': {'contentType': 'HTML', 'content': html},
                'toRecipients': [{'emailAddress': {'address': to_email}}],
            }
        },
        timeout=30,
    )

    if resp.status_code not in (200, 202):
        raise HTTPException(500, f'Failed to send email: {resp.status_code} {resp.text[:300]}')

    return {'ok': True, 'sent_to': to_email}

# ── Serve frontend ────────────────────────────────────────────────────────────
app.mount("/static", StaticFiles(directory=STATIC), name="static")

@app.get("/{full_path:path}")
def serve_spa(full_path: str):
    return FileResponse(
        os.path.join(STATIC, "index.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"}
    )
